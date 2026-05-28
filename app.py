import sys
import os
import logging
from datetime import datetime

from app_logging import APP_LOGGER_NAME, get_log_path, get_log_status, mark_logs_seen, setup_logging

setup_logging()
LOGGER = logging.getLogger(APP_LOGGER_NAME)

import gi

gi.require_version('Gtk', '4.0')
from gi.repository import Gtk, Gdk, GLib, Gio, Pango

import db
from models import UtilityService, UtilityCalculator
from constants import UKR_MONTHS, SERVICE_PREFIX_MAP, STATUS_LABELS
from ui.dialogs import ModalWindow
from ui.styles import CSS_STYLE
from ui.widgets import create_string_dropdown
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class UtilityTrackerApp(Gtk.Application):
    def __init__(self):
        super().__init__(application_id="io.github.f0ska.utilitytracker",
                         flags=Gio.ApplicationFlags.FLAGS_NONE)
        db.init_db()
        self.install_system_icon()
        self.active_tab = "dashboard"
        self.details_period = None
        self.details_origin = "dashboard"

    def install_system_icon(self):
        try:
            import shutil
            home = os.path.expanduser("~")
            dest_dir = os.path.join(home, ".local", "share", "icons", "hicolor", "scalable", "apps")
            os.makedirs(dest_dir, exist_ok=True)
            
            src_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icon.svg")
            if os.path.exists(src_path):
                shutil.copy2(src_path, os.path.join(dest_dir, "io.github.f0ska.utilitytracker.svg"))
                shutil.copy2(src_path, os.path.join(dest_dir, "icon.svg"))
        except Exception as e:
            LOGGER.warning("Could not install system icon", exc_info=True)

    def do_activate(self):
        self.win = Gtk.ApplicationWindow(application=self)
        self.win.set_title("UtilityTracker")
        self.win.set_default_size(950, 650)
        
        # Configure Window Icon for Taskbar/Dock
        try:
            self.win.set_icon_name("io.github.f0ska.utilitytracker")
        except Exception as e:
            LOGGER.warning("Could not set window icon", exc_info=True)
        
        # Load CSS
        css_provider = Gtk.CssProvider()
        css_provider.load_from_string(CSS_STYLE)
        Gtk.StyleContext.add_provider_for_display(
            Gdk.Display.get_default(),
            css_provider,
            Gtk.STYLE_PROVIDER_PRIORITY_APPLICATION
        )
        
        self.build_ui()
        self.win.present()

    # ==========================================
    # SHARED HELPERS
    # ==========================================
    @staticmethod
    def _clear_container(box):
        """Removes all children from a GTK Box container."""
        child = box.get_first_child()
        while child:
            next_child = child.get_next_sibling()
            box.remove(child)
            child = next_child

    @staticmethod
    def _format_period_name(period):
        """Converts 'YYYY-MM' into a localized Ukrainian string like 'Травень 2026'."""
        try:
            dt = datetime.strptime(period, "%Y-%m")
            return f"{UKR_MONTHS[dt.month]} {dt.year}"
        except Exception:
            return period

    def build_ui(self):
        # Main layout: Sidebar + Content Stack
        main_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        self.win.set_child(main_box)
        
        # Sidebar
        sidebar_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=5)
        sidebar_box.add_css_class("sidebar")
        sidebar_box.set_size_request(200, -1)
        main_box.append(sidebar_box)
        
        # Brand Title inside the sidebar
        title_label = Gtk.Label(label="UtilityTracker")
        title_label.add_css_class("title")
        title_label.set_margin_top(16)
        title_label.set_margin_bottom(16)
        title_label.set_halign(Gtk.Align.CENTER)
        sidebar_box.append(title_label)
        
        # Helper to create sidebar button with symbolic icon
        def create_sidebar_btn(icon_name, text_label):
            btn = Gtk.Button()
            btn.add_css_class("sidebar-button")
            
            box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
            box.set_halign(Gtk.Align.START)
            
            img = Gtk.Image.new_from_icon_name(icon_name)
            img.set_pixel_size(16)
            box.append(img)
            
            lbl = Gtk.Label(label=text_label)
            box.append(lbl)
            
            btn.set_child(box)
            return btn
        
        # Sidebar Navigation Buttons
        self.btn_dash = create_sidebar_btn("view-grid-symbolic", "Панель")
        self.btn_dash.add_css_class("sidebar-button-active")
        self.btn_dash.connect("clicked", lambda x: self.switch_tab("dashboard"))
        sidebar_box.append(self.btn_dash)
        
        self.btn_hist = create_sidebar_btn("document-open-recent-symbolic", "Історія")
        self.btn_hist.connect("clicked", lambda x: self.switch_tab("history"))
        sidebar_box.append(self.btn_hist)
        
        self.btn_tariffs = create_sidebar_btn("preferences-system-symbolic", "Тарифи")
        self.btn_tariffs.connect("clicked", lambda x: self.switch_tab("tariffs"))
        sidebar_box.append(self.btn_tariffs)
        
        self.btn_accs = create_sidebar_btn("view-list-symbolic", "Рахунки")
        self.btn_accs.connect("clicked", lambda x: self.switch_tab("accounts"))
        sidebar_box.append(self.btn_accs)
        
        # Spacer to push action button to the bottom (like LocalSend)
        spacer = Gtk.Box()
        spacer.set_vexpand(True)
        sidebar_box.append(spacer)
        
        # Action button at the bottom of the sidebar
        btn_add = Gtk.Button()
        btn_add.add_css_class("btn-primary")
        btn_add.set_margin_bottom(16)
        btn_add.set_margin_start(8)
        btn_add.set_margin_end(8)
        btn_add.connect("clicked", self.on_register_readings_clicked)
        
        box_add = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=8)
        box_add.set_halign(Gtk.Align.CENTER)
        img_add = Gtk.Image.new_from_icon_name("list-add-symbolic")
        img_add.set_pixel_size(16)
        box_add.append(img_add)
        lbl_add = Gtk.Label(label="Показники")
        box_add.append(lbl_add)
        btn_add.set_child(box_add)
        
        sidebar_box.append(btn_add)
        
        content_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
        content_box.set_hexpand(True)
        content_box.set_vexpand(True)
        main_box.append(content_box)

        self.log_banner = self.create_log_banner()
        content_box.append(self.log_banner)

        # Content Stack
        self.stack = Gtk.Stack()
        self.stack.set_transition_type(Gtk.StackTransitionType.SLIDE_LEFT_RIGHT)
        self.stack.set_transition_duration(250)
        self.stack.set_hexpand(True)
        self.stack.set_vexpand(True)
        content_box.append(self.stack)
        
        # Build views
        self.build_dashboard_view()
        self.build_history_view()
        self.build_tariffs_view()
        self.build_accounts_view()
        self.build_details_view()
        self.build_edit_history_view()
        
        # Initial refresh
        self.refresh_dashboard()
        self.refresh_log_banner()
        GLib.timeout_add_seconds(3, self.refresh_log_banner)

    def create_log_banner(self):
        banner = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
        banner.add_css_class("log-banner")
        banner.set_margin_start(16)
        banner.set_margin_end(16)
        banner.set_margin_top(8)
        banner.set_visible(False)

        icon = Gtk.Image.new_from_icon_name("dialog-warning-symbolic")
        icon.set_pixel_size(16)
        icon.set_valign(Gtk.Align.CENTER)
        banner.append(icon)

        self.log_banner_label = Gtk.Label()
        self.log_banner_label.set_hexpand(True)
        self.log_banner_label.set_halign(Gtk.Align.START)
        self.log_banner_label.set_valign(Gtk.Align.CENTER)
        self.log_banner_label.set_single_line_mode(True)
        self.log_banner_label.set_lines(1)
        self.log_banner_label.set_ellipsize(Pango.EllipsizeMode.END)
        banner.append(self.log_banner_label)

        btn_open = Gtk.Button(label="Відкрити лог")
        btn_open.add_css_class("btn-secondary")
        btn_open.set_valign(Gtk.Align.CENTER)
        btn_open.connect("clicked", self.on_open_log_clicked)
        banner.append(btn_open)

        btn_hide = Gtk.Button(label="Сховати")
        btn_hide.add_css_class("btn-secondary")
        btn_hide.set_valign(Gtk.Align.CENTER)
        btn_hide.connect("clicked", self.on_hide_log_banner_clicked)
        banner.append(btn_hide)

        return banner

    def refresh_log_banner(self):
        if not hasattr(self, "log_banner"):
            return True

        status = get_log_status()
        unseen_count = status["unseen_count"]
        self.log_banner.set_visible(unseen_count > 0)

        if unseen_count > 0:
            errors = status["error_count"]
            log_type = "помилки" if errors else "попередження"
            self.log_banner_label.set_text(
                f"Є нові {log_type} в журналі: {unseen_count}. Останній запис: {status['last_message']}"
            )

        return True

    def on_open_log_clicked(self, btn):
        log_path = get_log_path()
        try:
            Gio.AppInfo.launch_default_for_uri(
                GLib.filename_to_uri(log_path, None),
                None,
            )
        except Exception:
            self.show_info_dialog("Журнал", f"Файл журналу:\n{log_path}")

    def on_hide_log_banner_clicked(self, btn):
        mark_logs_seen()
        self.refresh_log_banner()

    def switch_tab(self, tab_name):
        self.active_tab = tab_name
        self.stack.set_visible_child_name(tab_name)
        
        # Update sidebar styling
        self.btn_dash.remove_css_class("sidebar-button-active")
        self.btn_hist.remove_css_class("sidebar-button-active")
        self.btn_tariffs.remove_css_class("sidebar-button-active")
        self.btn_accs.remove_css_class("sidebar-button-active")
        
        if tab_name == "dashboard":
            self.btn_dash.add_css_class("sidebar-button-active")
            self.refresh_dashboard()
        elif tab_name == "history":
            self.btn_hist.add_css_class("sidebar-button-active")
            self.refresh_history()
        elif tab_name == "tariffs":
            self.btn_tariffs.add_css_class("sidebar-button-active")
            self.refresh_tariffs()
        elif tab_name == "accounts":
            self.btn_accs.add_css_class("sidebar-button-active")
            self.refresh_accounts()

    # ==========================================
    # DASHBOARD VIEW
    # ==========================================
    def build_dashboard_view(self):
        scroll = Gtk.ScrolledWindow()
        scroll.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        
        self.dash_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        self.dash_box.add_css_class("dashboard-container")
        scroll.set_child(self.dash_box)
        
        self.stack.add_named(scroll, "dashboard")

    def refresh_dashboard(self):
        # Clear container
        self._clear_container(self.dash_box)
            
        # Get period (current month)
        now_period = datetime.now().strftime("%Y-%m")
        reading = db.get_reading(now_period)
        
        # If no reading for current, fallback to latest available month to display something
        if not reading:
            all_readings = db.get_all_readings()
            if all_readings:
                reading = all_readings[0]
                now_period = reading['period']
        
        # 1. Reminders
        reminders = UtilityService.get_reminders()
        for r in reminders:
            banner_class = "reminder-banner" if r['type'] == 'readings_missing' else "reminder-banner-info"
            banner = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=5)
            banner.add_css_class(banner_class)
            
            title = Gtk.Label(label=r['title'])
            title.set_halign(Gtk.Align.START)
            title.add_css_class("banner-title")
            banner.append(title)
            
            msg = Gtk.Label(label=r['message'])
            msg.set_halign(Gtk.Align.START)
            banner.append(msg)
            
            self.dash_box.append(banner)
            
        # 2. Main title for state
        title_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        title_box.set_valign(Gtk.Align.CENTER)
        
        lbl_period = Gtk.Label(label=f"Поточний стан на {self._format_period_name(now_period)}")
        lbl_period.set_halign(Gtk.Align.START)
        lbl_period.set_valign(Gtk.Align.CENTER)
        lbl_period.add_css_class("dashboard-section-title")
        lbl_period.set_hexpand(True)
        title_box.append(lbl_period)
        
        # Add share/report button
        if reading:
            btn_share = Gtk.Button(label="Швидкий звіт")
            btn_share.add_css_class("btn-secondary")
            btn_share.set_valign(Gtk.Align.CENTER)
            btn_share.connect("clicked", lambda x, p=now_period: self.show_report_modal(p))
            title_box.append(btn_share)
            
            btn_details = Gtk.Button(label="Детально")
            btn_details.add_css_class("btn-secondary")
            btn_details.set_valign(Gtk.Align.CENTER)
            btn_details.set_margin_start(6)
            btn_details.connect("clicked", lambda x, p=now_period: self.show_details_view(p, origin="dashboard"))
            title_box.append(btn_details)
            
        self.dash_box.append(title_box)
        
        # Add spacer after title box for nice spacing before grid
        title_spacer = Gtk.Box()
        title_spacer.set_size_request(-1, 8)
        self.dash_box.append(title_spacer)
        
        if not reading:
            no_data = Gtk.Label(label="Немає даних за цей період. Скористайтеся кнопкою «+ Показники» в бічній панелі ліворуч!")
            no_data.set_margin_top(40)
            self.dash_box.append(no_data)
            return
            
        # 3. Service Cards Grid (2 columns)
        grid = Gtk.Grid()
        grid.set_row_spacing(10)
        grid.set_column_spacing(10)
        grid.set_row_homogeneous(True)
        grid.set_column_homogeneous(True)
        self.dash_box.append(grid)
        
        # Load account numbers
        accounts = db.get_all_accounts()
        
        # Build cards for each service
        self.append_service_card(grid, 0, 0, now_period, "Електрика", "electricity", reading,
                                 f"Показники: День {reading['elec_day_reading']} | Ніч {reading['elec_night_reading']}",
                                 reading['elec_calculated'], reading['elec_billed'], reading['elec_paid_status'],
                                 accounts.get('electricity', ''))
                                 
        self.append_service_card(grid, 1, 0, now_period, "Газ (Споживання)", "gas", reading,
                                 f"Показник: {reading['gas_reading']}",
                                 reading['gas_calculated'], reading['gas_billed'], reading['gas_paid_status'],
                                 accounts.get('gas', ''))
                                 
        self.append_service_card(grid, 0, 1, now_period, "Газ (Розподіл)", "gas_dist", reading,
                                 f"Обсяг: {reading['gas_dist_volume']} м³",
                                 reading['gas_dist_calculated'], reading['gas_dist_billed'], reading['gas_dist_paid_status'],
                                 accounts.get('gas', ''))
                                 
        self.append_service_card(grid, 1, 1, now_period, "Водоканал", "water", reading,
                                 f"Показник: {reading['water_reading']}",
                                 reading['water_calculated'], reading['water_billed'], reading['water_paid_status'],
                                 accounts.get('water', ''))
                                 
        self.append_service_card(grid, 0, 2, now_period, "Вивіз сміття", "garbage", reading,
                                 "Фіксований тариф",
                                 reading['garbage_calculated'], reading['garbage_billed'], reading['garbage_paid_status'],
                                 accounts.get('garbage', ''))


    def append_service_card(self, grid, col, row, period, title, service_key, reading, readings_desc, calc_val, billed_val, status, account_num=""):
        card = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        card.add_css_class("card")
        
        # Header (Title + Status Badge)
        h_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        h_box.add_css_class("card-header")
        h_box.set_valign(Gtk.Align.CENTER)
        
        lbl_title = Gtk.Label(label=title)
        lbl_title.add_css_class("card-title")
        lbl_title.set_hexpand(True)
        lbl_title.set_halign(Gtk.Align.START)
        lbl_title.set_valign(Gtk.Align.CENTER)
        h_box.append(lbl_title)
        
        # Badge
        badge = Gtk.Label()
        badge.set_valign(Gtk.Align.CENTER)
        if status == 1:
            badge.set_label("Сплачено")
            badge.add_css_class("status-badge-paid")
        elif status == 2:
            badge.set_label("Аванс / Пропуск")
            badge.add_css_class("status-badge-prepaid")
        else:
            badge.set_label("Не сплачено")
            badge.add_css_class("status-badge-unpaid")
        h_box.append(badge)
        
        card.append(h_box)
        
        # Content
        lbl_desc = Gtk.Label(label=readings_desc)
        lbl_desc.set_halign(Gtk.Align.START)
        card.append(lbl_desc)
        
        # Account Number
        if account_num:
            lbl_acc = Gtk.Label(label=f"Особовий рахунок: {account_num}")
            lbl_acc.add_css_class("value-label")
            lbl_acc.set_halign(Gtk.Align.START)
            card.append(lbl_acc)
        
        # Calculated
        calc_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        lbl_c_title = Gtk.Label(label="Розраховано програмою:")
        lbl_c_title.add_css_class("value-label")
        lbl_c_val = Gtk.Label(label=f" {calc_val:.2f} грн")
        lbl_c_val.add_css_class("bold-value")
        calc_box.append(lbl_c_title)
        calc_box.append(lbl_c_val)
        card.append(calc_box)
        
        # Billed
        bill_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        lbl_b_title = Gtk.Label(label="Нараховано компанією:")
        lbl_b_title.add_css_class("value-label")
        bill_str = f" {billed_val:.2f} грн" if billed_val is not None else " -- грн"
        lbl_b_val = Gtk.Label(label=bill_str)
        lbl_b_val.add_css_class("bold-value")
        bill_box.append(lbl_b_title)
        bill_box.append(lbl_b_val)
        card.append(bill_box)
        
        # Paid
        db_prefix = SERVICE_PREFIX_MAP.get(service_key, service_key)
        paid_field = f"{db_prefix}_paid"
        paid_val = reading.get(paid_field, 0.0)
        paid_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        lbl_p_title = Gtk.Label(label="Сплачено фактично:")
        lbl_p_title.add_css_class("value-label")
        lbl_p_val = Gtk.Label(label=f" {paid_val:.2f} грн")
        lbl_p_val.add_css_class("bold-value")
        paid_box.append(lbl_p_title)
        paid_box.append(lbl_p_val)
        card.append(paid_box)
        
        # Buttons Box
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_margin_top(10)
        card.append(btn_box)
        
        # 1. Billed input button
        btn_bill = Gtk.Button(label="Нараховано")
        btn_bill.add_css_class("btn-secondary")
        btn_bill.connect("clicked", lambda x, p=period, s=service_key: self.on_enter_billed_clicked(p, s))
        btn_box.append(btn_bill)
        
        # 2. Pay Button (only if unpaid)
        if status == 0:
            btn_pay = Gtk.Button(label="Сплатити")
            btn_pay.add_css_class("btn-success")
            btn_pay.connect("clicked", lambda x, p=period, s=service_key: self.on_pay_clicked(p, s))
            btn_box.append(btn_pay)
            
            btn_skip = Gtk.Button(label="Аванс")
            btn_skip.add_css_class("btn-warning")
            btn_skip.connect("clicked", lambda x, p=period, s=service_key: self.on_skip_clicked(p, s))
            btn_box.append(btn_skip)
        else:
            # Edit payment button
            btn_edit_pay = Gtk.Button(label="Змінити оплату")
            btn_edit_pay.add_css_class("btn-secondary")
            btn_edit_pay.connect("clicked", lambda x, p=period, s=service_key: self.on_reset_payment_clicked(p, s))
            btn_box.append(btn_edit_pay)
            
        grid.attach(card, col, row, 1, 1)

    # ==========================================
    # HISTORY VIEW
    # ==========================================
    def build_history_view(self):
        scroll = Gtk.ScrolledWindow()
        scroll.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        
        self.hist_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        self.hist_box.add_css_class("dashboard-container")
        scroll.set_child(self.hist_box)
        
        self.stack.add_named(scroll, "history")

    def refresh_history(self):
        # Clear container
        self._clear_container(self.hist_box)
            
        header_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
        
        title = Gtk.Label(label="Історія платежів та показників")
        title.add_css_class("card-title")
        title.set_halign(Gtk.Align.START)
        title.set_hexpand(True)
        header_box.append(title)
        
        btn_export = Gtk.Button(label="Експорт")
        btn_export.add_css_class("btn-secondary")
        btn_export.set_valign(Gtk.Align.CENTER)
        btn_export.connect("clicked", self.on_export_clicked)
        header_box.append(btn_export)
        
        self.hist_box.append(header_box)
        
        all_readings = db.get_all_readings()
        if not all_readings:
            no_data = Gtk.Label(label="Історія порожня.")
            self.hist_box.append(no_data)
            return
            
        for r in all_readings:
            period = r['period']
            
            # Localize period name (Ukrainian months)
            period_name = self._format_period_name(period)
                
            # Compute total calculated and paid amounts
            tot_calc = (r.get('elec_calculated', 0.0) or 0.0) + \
                       (r.get('gas_calculated', 0.0) or 0.0) + \
                       (r.get('gas_dist_calculated', 0.0) or 0.0) + \
                       (r.get('water_calculated', 0.0) or 0.0) + \
                       (r.get('garbage_calculated', 0.0) or 0.0)
                       
            tot_paid = 0.0
            for f_paid, f_status in [('elec_paid', 'elec_paid_status'), ('gas_paid', 'gas_paid_status'),
                                     ('gas_dist_paid', 'gas_dist_paid_status'), ('water_paid', 'water_paid_status'),
                                     ('garbage_paid', 'garbage_paid_status')]:
                if r.get(f_status, 0) == 1:
                    tot_paid += (r.get(f_paid, 0.0) or 0.0)

            # Main card container (vertical box for structured rows)
            card = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=8)
            card.add_css_class("card")
            
            # 1. Top row: Period, Total Money Summary, and Buttons
            top_row = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
            top_row.set_valign(Gtk.Align.CENTER)
            card.append(top_row)
            
            lbl_p = Gtk.Label(label=period_name)
            lbl_p.add_css_class("card-title")
            lbl_p.set_hexpand(True)
            lbl_p.set_halign(Gtk.Align.START)
            top_row.append(lbl_p)
            
            lbl_totals = Gtk.Label()
            lbl_totals.set_markup(f"<span size='medium'>Сплачено: <span foreground='#15803d'><b>{tot_paid:.2f} грн</b></span></span>")
            lbl_totals.set_margin_end(15)
            lbl_totals.set_valign(Gtk.Align.CENTER)
            top_row.append(lbl_totals)
            
            btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
            top_row.append(btn_box)
            
            btn_details = Gtk.Button(label="Детально")
            btn_details.add_css_class("btn-secondary")
            btn_details.connect("clicked", lambda x, p=period: self.show_details_view(p, origin="history"))
            btn_box.append(btn_details)
            
            btn_edit = Gtk.Button(label="Редагувати")
            btn_edit.add_css_class("btn-secondary")
            btn_edit.connect("clicked", lambda x, p=period: self.show_edit_history_view(p))
            btn_box.append(btn_edit)
            
            btn_del = Gtk.Button(label="Видалити")
            btn_del.add_css_class("btn-danger")
            btn_del.connect("clicked", lambda x, p=period: self.on_delete_reading_clicked(p))
            btn_box.append(btn_del)
            
            # 2. Bottom row: Individual service paid details
            bottom_row = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=15)
            bottom_row.set_margin_top(4)
            card.append(bottom_row)
            
            def add_service_summary(box, name, paid, status):
                lbl = Gtk.Label()
                lbl.set_halign(Gtk.Align.START)
                if status == 1:
                    lbl.set_markup(f"<span size='small' foreground='#15803d'><b>{name}:</b> {paid:.2f} грн</span>")
                elif status == 2:
                    lbl.set_markup(f"<span size='small' foreground='#3b82f6'><b>{name}:</b> Аванс</span>")
                else:
                    lbl.set_markup(f"<span size='small' foreground='#ef4444'><b>{name}:</b> Не сплачено</span>")
                box.append(lbl)

            if r.get('elec_day_reading') is not None:
                add_service_summary(bottom_row, "Електрика", r.get('elec_paid', 0.0), r.get('elec_paid_status', 0))
            if r.get('gas_reading') is not None:
                add_service_summary(bottom_row, "Газ", r.get('gas_paid', 0.0), r.get('gas_paid_status', 0))
            if r.get('gas_dist_volume') is not None:
                add_service_summary(bottom_row, "Розподіл", r.get('gas_dist_paid', 0.0), r.get('gas_dist_paid_status', 0))
            if r.get('water_reading') is not None:
                add_service_summary(bottom_row, "Вода", r.get('water_paid', 0.0), r.get('water_paid_status', 0))
                
            add_service_summary(bottom_row, "Сміття", r.get('garbage_paid', 0.0), r.get('garbage_paid_status', 0))
            
            self.hist_box.append(card)

    # ==========================================
    # TARIFFS VIEW
    # ==========================================
    def build_tariffs_view(self):
        scroll = Gtk.ScrolledWindow()
        scroll.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        
        self.tariffs_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        self.tariffs_box.add_css_class("dashboard-container")
        scroll.set_child(self.tariffs_box)
        
        self.stack.add_named(scroll, "tariffs")

    def refresh_tariffs(self):
        # Clear container
        self._clear_container(self.tariffs_box)
            
        title = Gtk.Label(label="Налаштування тарифів")
        title.add_css_class("card-title")
        title.set_halign(Gtk.Align.START)
        self.tariffs_box.append(title)
        
        tariffs = db.get_all_tariffs()
        
        # Grid of inputs
        grid = Gtk.Grid()
        grid.set_row_spacing(10)
        grid.set_column_spacing(15)
        self.tariffs_box.append(grid)
        
        self.tariff_entries = {}
        
        tariff_labels = {
            'electricity_day': 'Електрика День (грн/кВт):',
            'electricity_night': 'Електрика Ніч (грн/кВт):',
            'gas': 'Газ Споживання (грн/м³):',
            'gas_distribution': 'Газ Розподіл Тариф (грн):',
            'gas_distribution_volume': 'Газ Розподіл Річний Обсяг (м³):',
            'water_supply': 'Водопостачання Тариф (грн/м³):',
            'water_drainage': 'Водовідведення Тариф (грн/м³):',
            'water_subscription': 'Вода Абонплата (грн):',
            'garbage': 'Вивіз сміття Тариф (грн):'
        }
        
        idx = 0
        for key, label_text in tariff_labels.items():
            lbl = Gtk.Label(label=label_text)
            lbl.set_halign(Gtk.Align.START)
            grid.attach(lbl, 0, idx, 1, 1)
            
            val = tariffs.get(key, 0.0)
            entry = Gtk.Entry()
            entry.set_text(str(val))
            grid.attach(entry, 1, idx, 1, 1)
            
            self.tariff_entries[key] = entry
            idx += 1
            
        btn_save = Gtk.Button(label="Зберегти тарифи")
        btn_save.add_css_class("btn-primary")
        btn_save.connect("clicked", self.on_save_tariffs_clicked)
        self.tariffs_box.append(btn_save)

    def on_save_tariffs_clicked(self, btn):
        tariffs = db.get_all_tariffs()
        changed = []
        
        for key, entry in self.tariff_entries.items():
            try:
                new_val = float(entry.get_text())
                if abs(new_val - tariffs.get(key, 0.0)) > 1e-6:
                    changed.append((key, new_val))
            except ValueError:
                # Invalid float
                self.show_error_dialog("Помилка", f"Введено некоректне значення для тарифу: {key}")
                return
                
        if not changed:
            self.show_info_dialog("Інформація", "Тарифи не змінилися.")
            return
            
        # Check if we have unpaid readings to offer recalculation
        unpaid = db.get_unpaid_readings()
        if unpaid:
            # Show recalculation prompt
            self.show_recalc_prompt_dialog(changed)
        else:
            # Just save
            for key, val in changed:
                UtilityService.update_single_tariff_and_recalculate(key, val, recalculate_unpaid=False)
            self.show_info_dialog("Успішно", "Тарифи збережено!")
            self.refresh_tariffs()

    def show_recalc_prompt_dialog(self, changed_tariffs):
        dialog = ModalWindow(title="Зміна тарифів", transient_for=self.win, modal=True)
        dialog.set_default_size(420, -1)
        
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        box.set_margin_start(15)
        box.set_margin_end(15)
        content.append(box)
        
        lbl = Gtk.Label(label="Ви змінили тарифи. Бажаєте перерахувати суми до сплати для неоплачених періодів за новими тарифами?")
        lbl.set_wrap(True)
        lbl.set_halign(Gtk.Align.START)
        box.append(lbl)
        
        # Left-aligned button container
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        
        btn_yes = Gtk.Button(label="Так, перерахувати")
        btn_yes.add_css_class("btn-success")
        
        btn_no = Gtk.Button(label="Ні, тільки зберегти")
        btn_no.add_css_class("btn-secondary")
        
        btn_cancel = Gtk.Button(label="Скасувати")
        btn_cancel.add_css_class("btn-secondary")
        
        btn_box.append(btn_yes)
        btn_box.append(btn_no)
        btn_box.append(btn_cancel)
        box.append(btn_box)
        
        def do_yes(btn):
            for key, val in changed_tariffs:
                UtilityService.update_single_tariff_and_recalculate(key, val, recalculate_unpaid=True)
            self.show_info_dialog("Успішно", "Тарифи збережено та неоплачені рахунки перераховано!")
            dialog.destroy()
            self.refresh_tariffs()
            
        def do_no(btn):
            for key, val in changed_tariffs:
                UtilityService.update_single_tariff_and_recalculate(key, val, recalculate_unpaid=False)
            self.show_info_dialog("Успішно", "Тарифи збережено без перерахунку.")
            dialog.destroy()
            self.refresh_tariffs()
            
        btn_yes.connect("clicked", do_yes)
        btn_no.connect("clicked", do_no)
        btn_cancel.connect("clicked", lambda x: dialog.destroy())
        
        dialog.present()

    # ==========================================
    # ACTION DIALOGS & HANDLERS
    # ==========================================
    def on_register_readings_clicked(self, btn):
        self.show_readings_entry_dialog(None)

    def on_edit_reading_clicked(self, period):
        self.show_readings_entry_dialog(period)

    def show_readings_entry_dialog(self, period=None):
        title_text = "Редагування показників" if period else "Зареєструвати показники"
        dialog = ModalWindow(title=title_text, transient_for=self.win, modal=True)
        dialog.set_default_size(450, -1)
        
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=12)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        box.set_margin_start(15)
        box.set_margin_end(15)
        content.append(box)
        
        # Period selection
        p_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
        p_lbl = Gtk.Label(label="Період (РРРР-ММ):")
        p_entry = Gtk.Entry()
        if period:
            p_entry.set_text(period)
            p_entry.set_sensitive(False) # Cannot edit period of existing
        else:
            p_entry.set_text(datetime.now().strftime("%Y-%m"))
        p_box.append(p_lbl)
        p_box.append(p_entry)
        box.append(p_box)
        
        # Load existing data if edit mode
        existing = db.get_reading(period) if period else {}
        
        # Entry rows helper
        entries = {}
        def add_row(label_text, key, val_type=float):
            r_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
            lbl = Gtk.Label(label=label_text)
            lbl.set_hexpand(True)
            lbl.set_halign(Gtk.Align.START)
            ent = Gtk.Entry()
            
            curr_val = existing.get(key)
            if curr_val is not None:
                ent.set_text(str(curr_val))
            elif key == "gas_dist_volume":
                tariffs = db.get_all_tariffs()
                curr_val = tariffs.get("gas_distribution_volume", 70.66)
                ent.set_text(str(curr_val))
                
            r_box.append(lbl)
            r_box.append(ent)
            box.append(r_box)
            entries[key] = ent
            
        add_row("Електрика День (кВт):", "elec_day_reading")
        add_row("Електрика Ніч (кВт):", "elec_night_reading")
        add_row("Газ Показник (м³):", "gas_reading")
        add_row("Газ Розподіл Обсяг (м³):", "gas_dist_volume")
        add_row("Вода Показник (м³):", "water_reading")
        
        # Left-aligned button container
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        
        btn_save = Gtk.Button(label="Зберегти")
        btn_save.add_css_class("btn-primary")
        
        btn_cancel = Gtk.Button(label="Скасувати")
        btn_cancel.add_css_class("btn-secondary")
        
        btn_box.append(btn_save)
        btn_box.append(btn_cancel)
        box.append(btn_box)
        
        def do_save(btn):
            p_text = p_entry.get_text().strip()
            if not p_text:
                self.show_error_dialog("Помилка", "Період не може бути порожнім!")
                return
                
            try:
                e_day = float(entries["elec_day_reading"].get_text()) if entries["elec_day_reading"].get_text() else None
                e_night = float(entries["elec_night_reading"].get_text()) if entries["elec_night_reading"].get_text() else None
                g_read = float(entries["gas_reading"].get_text()) if entries["gas_reading"].get_text() else None
                g_dist = float(entries["gas_dist_volume"].get_text()) if entries["gas_dist_volume"].get_text() else None
                w_read = float(entries["water_reading"].get_text()) if entries["water_reading"].get_text() else None
                
                UtilityService.register_readings(p_text, e_day, e_night, g_read, w_read, g_dist)
                
                self.refresh_dashboard()
                self.refresh_history()
                dialog.destroy()
            except ValueError:
                self.show_error_dialog("Помилка", "Будь ласка, введіть числові значення показників!")
                
        btn_save.connect("clicked", do_save)
        btn_cancel.connect("clicked", lambda x: dialog.destroy())
        
        dialog.present()

    def on_delete_reading_clicked(self, period):
        dialog = ModalWindow(title="Видалення", transient_for=self.win, modal=True)
        dialog.set_default_size(380, -1)
        
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        box.set_margin_start(15)
        box.set_margin_end(15)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        content.append(box)
        
        lbl = Gtk.Label(label=f"Ви впевнені, що хочете видалити запис за період {period}?")
        lbl.set_wrap(True)
        lbl.set_halign(Gtk.Align.START)
        box.append(lbl)
        
        # Left-aligned button container
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        
        btn_del = Gtk.Button(label="Видалити")
        btn_del.add_css_class("btn-warning")
        
        btn_cancel = Gtk.Button(label="Скасувати")
        btn_cancel.add_css_class("btn-secondary")
        
        btn_box.append(btn_del)
        btn_box.append(btn_cancel)
        box.append(btn_box)
        
        def do_delete(btn):
            db.delete_reading(period)
            self.refresh_dashboard()
            self.refresh_history()
            dialog.destroy()
            
        btn_del.connect("clicked", do_delete)
        btn_cancel.connect("clicked", lambda x: dialog.destroy())
        
        dialog.present()

    def on_enter_billed_clicked(self, period, service):
        dialog = ModalWindow(title="Внесення нарахування", transient_for=self.win, modal=True)
        dialog.set_default_size(350, -1)
        
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        box.set_margin_start(15)
        box.set_margin_end(15)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        content.append(box)
        
        lbl = Gtk.Label(label=f"Введіть суму, нараховану компанією:")
        box.append(lbl)
        
        entry = Gtk.Entry()
        # Prepopulate existing
        record = db.get_reading(period)
        prefix = SERVICE_PREFIX_MAP.get(service)
        existing_val = record.get(f"{prefix}_billed") if record else None
        if existing_val is not None:
            entry.set_text(str(existing_val))
        box.append(entry)
        
        # Left-aligned button container
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        
        btn_save = Gtk.Button(label="Зберегти")
        btn_save.add_css_class("btn-primary")
        
        btn_cancel = Gtk.Button(label="Скасувати")
        btn_cancel.add_css_class("btn-secondary")
        
        btn_box.append(btn_save)
        btn_box.append(btn_cancel)
        box.append(btn_box)
        
        def do_save(btn):
            try:
                val = float(entry.get_text())
                record[f"{prefix}_billed"] = val
                db.save_reading(record)
                self.refresh_dashboard()
                dialog.destroy()
            except ValueError:
                self.show_error_dialog("Помилка", "Будь ласка, введіть числове значення!")
                
        btn_save.connect("clicked", do_save)
        btn_cancel.connect("clicked", lambda x: dialog.destroy())
        
        dialog.present()

    def on_pay_clicked(self, period, service):
        record = db.get_reading(period)
        if not record:
            return
            
        prefix = SERVICE_PREFIX_MAP.get(service)
        
        calc_val = record.get(f"{prefix}_calculated") or 0.0
        billed_val = record.get(f"{prefix}_billed")
        
        dialog = ModalWindow(title="Сплатити", transient_for=self.win, modal=True)
        dialog.set_default_size(400, -1)
        
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=12)
        box.set_margin_start(15)
        box.set_margin_end(15)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        content.append(box)
        
        lbl = Gtk.Label(label="Виберіть суму для оплати:")
        lbl.add_css_class("card-title")
        box.append(lbl)
        
        # Radio buttons (simulated with standard check buttons as group)
        rad_calc = Gtk.CheckButton(label=f"Розрахована програмою: {calc_val:.2f} грн")
        rad_calc.set_active(True)
        box.append(rad_calc)
        
        rad_billed = Gtk.CheckButton(label="Нарахована компанією: -- грн")
        rad_billed.set_group(rad_calc)
        if billed_val is not None:
            rad_billed.set_label(f"Нарахована компанією: {billed_val:.2f} грн")
        else:
            rad_billed.set_sensitive(False)
        box.append(rad_billed)
        
        rad_custom = Gtk.CheckButton(label="Інша сума (ввести вручну):")
        rad_custom.set_group(rad_calc)
        box.append(rad_custom)
        
        custom_entry = Gtk.Entry()
        custom_entry.set_placeholder_text("Введіть суму...")
        custom_entry.set_sensitive(False)
        box.append(custom_entry)
        
        # Enable entry only when custom is selected
        def on_rad_toggled(btn):
            custom_entry.set_sensitive(rad_custom.get_active())
        rad_custom.connect("toggled", on_rad_toggled)
        
        # Left-aligned button container
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        
        btn_submit = Gtk.Button(label="Підтвердити оплату")
        btn_submit.add_css_class("btn-primary")
        
        btn_cancel = Gtk.Button(label="Скасувати")
        btn_cancel.add_css_class("btn-secondary")
        
        btn_box.append(btn_submit)
        btn_box.append(btn_cancel)
        box.append(btn_box)
        
        def do_submit(btn):
            choice = 'calculated'
            custom_val = None
            
            if rad_billed.get_active():
                choice = 'billed'
            elif rad_custom.get_active():
                choice = 'custom'
                try:
                    custom_val = float(custom_entry.get_text())
                except ValueError:
                    self.show_error_dialog("Помилка", "Будь ласка, введіть числове значення суми!")
                    return
                    
            UtilityService.mark_as_paid(period, service, choice, custom_val)
            self.refresh_dashboard()
            dialog.destroy()
            
        btn_submit.connect("clicked", do_submit)
        btn_cancel.connect("clicked", lambda x: dialog.destroy())
        
        dialog.present()

    def on_skip_clicked(self, period, service):
        UtilityService.mark_as_prepaid_or_skip(period, service)
        self.show_info_dialog("Успішно", "Позначено як Аванс / Пропущено. Ця послуга більше не вважатиметься боргом за цей місяць.")
        self.refresh_dashboard()

    def on_reset_payment_clicked(self, period, service):
        record = db.get_reading(period)
        if not record:
            return
            
        prefix = SERVICE_PREFIX_MAP.get(service)
        
        record[f"{prefix}_paid"] = 0.0
        record[f"{prefix}_paid_status"] = 0 # Unpaid
        record[f"{prefix}_paid_date"] = None
        db.save_reading(record)
        self.refresh_dashboard()

    def show_report_modal(self, period):
        reading = db.get_reading(period)
        if not reading:
            return
            
        accounts = db.get_all_accounts()
        
        dialog = ModalWindow(title="Звіт для месенджерів", transient_for=self.win, modal=True)
        dialog.set_default_size(550, 690)
        
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        box.set_margin_start(20)
        box.set_margin_end(20)
        box.set_margin_top(20)
        box.set_margin_bottom(20)
        content.append(box)
        
        title = Gtk.Label(label="Передача показників в один клік")
        title.add_css_class("dashboard-section-title")
        title.set_halign(Gtk.Align.START)
        box.append(title)
        
        # Scrollable container for cards
        scroll = Gtk.ScrolledWindow()
        scroll.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        scroll.set_vexpand(True)
        box.append(scroll)
        
        list_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        scroll.set_child(list_box)
        
        # Helper to create a unified service card in the modal
        def create_service_card(card_title_text):
            card_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
            card_box.add_css_class("card")
            
            lbl_title = Gtk.Label(label=card_title_text)
            lbl_title.add_css_class("card-title")
            lbl_title.set_halign(Gtk.Align.START)
            card_box.append(lbl_title)
            
            list_box.append(card_box)
            return card_box

        # Helper to add copyable item rows inside a service card
        def add_copy_item(card_box, label_text, display_text, copy_data):
            item_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=4)
            item_box.set_margin_bottom(6)
            
            if label_text:
                lbl_label = Gtk.Label(label=label_text)
                lbl_label.add_css_class("value-label")
                lbl_label.set_halign(Gtk.Align.START)
                item_box.append(lbl_label)
                
            h_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
            h_box.set_valign(Gtk.Align.CENTER)
            item_box.append(h_box)
            
            lbl_text = Gtk.Label(label=display_text)
            lbl_text.set_hexpand(True)
            lbl_text.set_halign(Gtk.Align.START)
            lbl_text.set_selectable(True)
            h_box.append(lbl_text)
            
            btn_copy = Gtk.Button(label="Копіювати")
            btn_copy.add_css_class("btn-secondary")
            btn_copy.set_valign(Gtk.Align.CENTER)
            
            def do_copy(btn):
                clipboard = self.win.get_clipboard()
                clipboard.set(str(copy_data))
                # Micro-animation
                btn.set_label("Скопіювано! ✓")
                GLib.timeout_add(1500, lambda: btn.set_label("Копіювати") or False)
                
            btn_copy.connect("clicked", do_copy)
            h_box.append(btn_copy)
            
            card_box.append(item_box)

        # 1. Electricity Card (single copy row)
        elec_acc = accounts.get('electricity', '')
        e_night = reading.get('elec_night_reading')
        e_day = reading.get('elec_day_reading')
        if e_night is not None and e_day is not None:
            elec_card = create_service_card("Електрика")
            elec_text = f"{elec_acc} {e_night:.0f} {e_day:.0f}".strip()
            add_copy_item(elec_card, "Реквізити + Показники (Ніч/День):", elec_text, elec_text)

        # 2. Gas Card (combined copy rows)
        gas_acc = accounts.get('gas', '')
        gas_read = reading.get('gas_reading')
        if gas_acc or gas_read is not None:
            gas_card = create_service_card("Газ")
            if gas_acc:
                add_copy_item(gas_card, "Особовий рахунок розподілу:", gas_acc, gas_acc)
            if gas_read is not None:
                gas_read_str = f"{gas_read:.0f}"
                add_copy_item(gas_card, "Показник лічильника споживання:", gas_read_str, gas_read_str)

        # 3. Water Card (combined copy rows)
        water_acc = accounts.get('water', '')
        water_read = reading.get('water_reading')
        if water_acc or water_read is not None:
            water_card = create_service_card("Водоканал")
            if water_acc:
                add_copy_item(water_card, "Особовий рахунок:", water_acc, water_acc)
            if water_read is not None:
                water_read_str = f"{water_read:.0f}"
                add_copy_item(water_card, "Показник лічильника:", water_read_str, water_read_str)

        dialog.present()

    # ==========================================
    # UTILITY DIALOGS
    # ==========================================
    def show_info_dialog(self, title, msg):
        dialog = ModalWindow(title=title, transient_for=self.win, modal=True)
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        box.set_margin_start(15)
        box.set_margin_end(15)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        content.append(box)
        
        lbl = Gtk.Label(label=msg)
        lbl.set_wrap(True)
        lbl.set_halign(Gtk.Align.START)
        box.append(lbl)
        
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        
        btn_ok = Gtk.Button(label="OK")
        btn_ok.add_css_class("btn-secondary")
        btn_ok.connect("clicked", lambda x: dialog.destroy())
        btn_box.append(btn_ok)
        box.append(btn_box)
        
        dialog.present()

    def show_error_dialog(self, title, msg):
        LOGGER.error("%s: %s", title, msg)
        self.refresh_log_banner()
        dialog = ModalWindow(title=title, transient_for=self.win, modal=True)
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        box.set_margin_start(15)
        box.set_margin_end(15)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        content.append(box)
        
        lbl = Gtk.Label(label=f"❌ {msg}")
        lbl.set_wrap(True)
        lbl.set_halign(Gtk.Align.START)
        box.append(lbl)
        
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        
        btn_ok = Gtk.Button(label="OK")
        btn_ok.add_css_class("btn-secondary")
        btn_ok.connect("clicked", lambda x: dialog.destroy())
        btn_box.append(btn_ok)
        box.append(btn_box)
        
        dialog.present()

    # ==========================================
    # ACCOUNTS VIEW
    # ==========================================
    def build_accounts_view(self):
        scroll = Gtk.ScrolledWindow()
        scroll.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        
        self.accounts_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        self.accounts_box.add_css_class("dashboard-container")
        scroll.set_child(self.accounts_box)
        
        self.stack.add_named(scroll, "accounts")

    def refresh_accounts(self):
        # Clear container
        self._clear_container(self.accounts_box)
            
        title = Gtk.Label(label="Особові рахунки")
        title.add_css_class("card-title")
        title.set_halign(Gtk.Align.START)
        self.accounts_box.append(title)
        
        accounts = db.get_all_accounts()
        
        grid = Gtk.Grid()
        grid.set_row_spacing(10)
        grid.set_column_spacing(15)
        self.accounts_box.append(grid)
        
        self.account_entries = {}
        
        account_labels = {
            'electricity': 'Електрика (Особовий рахунок):',
            'gas': 'Газ (Особовий рахунок):',
            'water': 'Вода (Особовий рахунок):',
            'garbage': 'Вивіз сміття (Особовий рахунок):'
        }
        
        idx = 0
        for key, label_text in account_labels.items():
            lbl = Gtk.Label(label=label_text)
            lbl.set_halign(Gtk.Align.START)
            grid.attach(lbl, 0, idx, 1, 1)
            
            val = accounts.get(key, '')
            entry = Gtk.Entry()
            entry.set_text(str(val))
            entry.set_hexpand(True) # Automatically expand to the full width
            grid.attach(entry, 1, idx, 1, 1)
            
            self.account_entries[key] = entry
            idx += 1
            
        btn_save = Gtk.Button(label="Зберегти рахунки")
        btn_save.add_css_class("btn-primary")
        btn_save.connect("clicked", self.on_save_accounts_clicked)
        self.accounts_box.append(btn_save)

    def on_save_accounts_clicked(self, btn):
        for key, entry in self.account_entries.items():
            val = entry.get_text().strip()
            db.update_account(key, val)
            
        self.show_info_dialog("Успішно", "Особові рахунки збережено!")
        self.refresh_accounts()
        self.refresh_dashboard()

    def build_details_view(self):
        scroll = Gtk.ScrolledWindow()
        scroll.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        
        self.details_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        self.details_box.add_css_class("dashboard-container")
        scroll.set_child(self.details_box)
        
        self.stack.add_named(scroll, "details")

    def show_details_view(self, period, origin="dashboard"):
        self.details_period = period
        self.details_origin = origin
        
        # Switch tab directly and deactivate sidebar buttons
        self.active_tab = "details"
        self.stack.set_visible_child_name("details")
        
        self.btn_dash.remove_css_class("sidebar-button-active")
        self.btn_hist.remove_css_class("sidebar-button-active")
        self.btn_tariffs.remove_css_class("sidebar-button-active")
        self.btn_accs.remove_css_class("sidebar-button-active")
        
        # Clear previous elements from container
        self._clear_container(self.details_box)
            
        record = db.get_reading(period)
        if not record:
            self.show_error_dialog("Помилка", f"Не знайдено даних для періоду {period}!")
            self.switch_tab(origin)
            return
            
        prev_record = db.get_previous_reading(period) or {}
        
        # Header navigation row
        header_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=15)
        header_box.add_css_class("details-header-box")
        
        btn_back = Gtk.Button(label="← Назад")
        btn_back.add_css_class("btn-secondary")
        btn_back.connect("clicked", lambda x: self.switch_tab(self.details_origin))
        header_box.append(btn_back)
        
        # Format month name nicely using shared helper
        title_text = f"Деталі розрахунків: {self._format_period_name(period)}"
            
        lbl_title = Gtk.Label(label=title_text)
        lbl_title.add_css_class("dashboard-section-title")
        lbl_title.set_hexpand(True)
        lbl_title.set_halign(Gtk.Align.START)
        header_box.append(lbl_title)
        
        self.details_box.append(header_box)
        
        notebook = Gtk.Notebook()
        self.details_box.append(notebook)
        
        def make_cell(text, is_header=False, is_bold=False, halign=Gtk.Align.START):
            lbl = Gtk.Label(label=str(text))
            lbl.set_halign(halign)
            lbl.set_hexpand(True)
            if is_header:
                lbl.add_css_class("details-grid-header")
            elif is_bold:
                lbl.add_css_class("details-grid-cell-bold")
            else:
                lbl.add_css_class("details-grid-cell")
            return lbl

        def get_status_text(status):
            return STATUS_LABELS.get(status, STATUS_LABELS[0])

        # ----------------------------------------------------
        # TAB 1: Зведена таблиця (Master Summary Table)
        # ----------------------------------------------------
        tab_summary = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=10)
        tab_summary.set_margin_start(10)
        tab_summary.set_margin_end(10)
        tab_summary.set_margin_top(15)
        tab_summary.set_margin_bottom(10)
        
        lbl_sum_title = Gtk.Label(label="Зведена таблиця розрахунків")
        lbl_sum_title.add_css_class("details-tab-title")
        lbl_sum_title.set_halign(Gtk.Align.START)
        tab_summary.append(lbl_sum_title)
        
        grid_sum = Gtk.Grid()
        grid_sum.add_css_class("details-grid")
        grid_sum.set_row_spacing(2)
        grid_sum.set_column_spacing(2)
        tab_summary.append(grid_sum)
        
        headers = ["Послуга", "Показники / Обсяг", "Тариф(и)", "Розраховано", "Нараховано", "Сплачено", "Статус"]
        for col_idx, h_text in enumerate(headers):
            grid_sum.attach(make_cell(h_text, is_header=True), col_idx, 0, 1, 1)
            
        rows_data = []
        # Electricity Row
        e_curr_d = record.get('elec_day_reading')
        e_prev_d = prev_record.get('elec_day_reading', e_curr_d)
        e_curr_n = record.get('elec_night_reading')
        e_prev_n = prev_record.get('elec_night_reading', e_curr_n)
        
        e_vol_str = "-"
        if e_curr_d is not None and e_curr_n is not None and e_prev_d is not None and e_prev_n is not None:
            e_vol_str = f"День: {e_curr_d - e_prev_d:.0f} кВт\nНіч: {e_curr_n - e_prev_n:.0f} кВт"
        e_tariff_str = f"День: {record.get('elec_day_tariff', 0.0):.2f}\nНіч: {record.get('elec_night_tariff', 0.0):.2f}"
        
        rows_data.append((
            "Електрика", e_vol_str, e_tariff_str,
            f"{record.get('elec_calculated', 0.0):.2f} грн",
            f"{record.get('elec_billed'):.2f} грн" if record.get('elec_billed') is not None else "-",
            f"{record.get('elec_paid', 0.0):.2f} грн" if record.get('elec_paid_status', 0) == 1 else "-",
            get_status_text(record.get('elec_paid_status', 0))
        ))
        
        # Gas consumption
        g_curr = record.get('gas_reading')
        g_prev = prev_record.get('gas_reading', g_curr)
        g_vol_str = f"{g_curr - g_prev:.2f} м³" if g_curr is not None and g_prev is not None else "-"
        rows_data.append((
            "Газ (Споживання)", g_vol_str, f"{record.get('gas_tariff', 0.0):.4f}",
            f"{record.get('gas_calculated', 0.0):.2f} грн",
            f"{record.get('gas_billed'):.2f} грн" if record.get('gas_billed') is not None else "-",
            f"{record.get('gas_paid', 0.0):.2f} грн" if record.get('gas_paid_status', 0) == 1 else "-",
            get_status_text(record.get('gas_paid_status', 0))
        ))
        
        # Gas distribution
        rows_data.append((
            "Газ (Розподіл)", f"{record.get('gas_dist_volume', 0.0):.2f} м³", f"{record.get('gas_dist_tariff', 0.0):.4f}",
            f"{record.get('gas_dist_calculated', 0.0):.2f} грн",
            f"{record.get('gas_dist_billed'):.2f} грн" if record.get('gas_dist_billed') is not None else "-",
            f"{record.get('gas_dist_paid', 0.0):.2f} грн" if record.get('gas_dist_paid_status', 0) == 1 else "-",
            get_status_text(record.get('gas_dist_paid_status', 0))
        ))
        
        # Water Row
        w_curr = record.get('water_reading')
        w_prev = prev_record.get('water_reading', w_curr)
        w_vol_str = f"{w_curr - w_prev:.2f} м³" if w_curr is not None and w_prev is not None else "-"
        w_tariff_str = f"Пост: {record.get('water_supply_tariff', 0.0):.2f}\nВідв: {record.get('water_drainage_tariff', 0.0):.2f}\nАбон: {record.get('water_sub_tariff', 0.0):.2f}"
        rows_data.append((
            "Водоканал", w_vol_str, w_tariff_str,
            f"{record.get('water_calculated', 0.0):.2f} грн",
            f"{record.get('water_billed'):.2f} грн" if record.get('water_billed') is not None else "-",
            f"{record.get('water_paid', 0.0):.2f} грн" if record.get('water_paid_status', 0) == 1 else "-",
            get_status_text(record.get('water_paid_status', 0))
        ))
        
        # Garbage Row
        rows_data.append((
            "Вивіз сміття", "Фіксований тариф", f"{record.get('garbage_tariff', 0.0):.2f}",
            f"{record.get('garbage_calculated', 0.0):.2f} грн",
            f"{record.get('garbage_billed'):.2f} грн" if record.get('garbage_billed') is not None else "-",
            f"{record.get('garbage_paid', 0.0):.2f} грн" if record.get('garbage_paid_status', 0) == 1 else "-",
            get_status_text(record.get('garbage_paid_status', 0))
        ))
        
        # Populate Grid Rows
        for row_idx, row_values in enumerate(rows_data, start=1):
            for col_idx, val in enumerate(row_values):
                grid_sum.attach(make_cell(val), col_idx, row_idx, 1, 1)
                
        # Total sums row
        tot_calc = (record.get('elec_calculated', 0.0) or 0.0) + \
                   (record.get('gas_calculated', 0.0) or 0.0) + \
                   (record.get('gas_dist_calculated', 0.0) or 0.0) + \
                   (record.get('water_calculated', 0.0) or 0.0) + \
                   (record.get('garbage_calculated', 0.0) or 0.0)
                   
        tot_billed = 0.0
        has_billed = False
        for f in ['elec_billed', 'gas_billed', 'gas_dist_billed', 'water_billed', 'garbage_billed']:
            if record.get(f) is not None:
                tot_billed += record.get(f)
                has_billed = True
                
        tot_paid = 0.0
        for f_paid, f_status in [('elec_paid', 'elec_paid_status'), ('gas_paid', 'gas_paid_status'),
                                 ('gas_dist_paid', 'gas_dist_paid_status'), ('water_paid', 'water_paid_status'),
                                 ('garbage_paid', 'garbage_paid_status')]:
            if record.get(f_status, 0) == 1:
                tot_paid += (record.get(f_paid, 0.0) or 0.0)
                
        last_row_idx = len(rows_data) + 1
        grid_sum.attach(make_cell("РАЗОМ", is_bold=True), 0, last_row_idx, 1, 1)
        grid_sum.attach(make_cell(""), 1, last_row_idx, 1, 1)
        grid_sum.attach(make_cell(""), 2, last_row_idx, 1, 1)
        grid_sum.attach(make_cell(f"{tot_calc:.2f} грн", is_bold=True), 3, last_row_idx, 1, 1)
        grid_sum.attach(make_cell(f"{tot_billed:.2f} грн" if has_billed else "-", is_bold=True), 4, last_row_idx, 1, 1)
        grid_sum.attach(make_cell(f"{tot_paid:.2f} грн", is_bold=True), 5, last_row_idx, 1, 1)
        grid_sum.attach(make_cell(""), 6, last_row_idx, 1, 1)
        
        notebook.append_page(tab_summary, Gtk.Label(label="Зведена"))

        # ----------------------------------------------------
        # TAB 2: Електрика (Electricity Details)
        # ----------------------------------------------------
        tab_elec = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        tab_elec.set_margin_start(10)
        tab_elec.set_margin_end(10)
        tab_elec.set_margin_top(15)
        tab_elec.set_margin_bottom(10)
        
        lbl_e_title = Gtk.Label(label="Деталі розрахунку за електроенергію (День/Ніч)")
        lbl_e_title.add_css_class("details-tab-title")
        lbl_e_title.set_halign(Gtk.Align.START)
        tab_elec.append(lbl_e_title)
        
        grid_elec = Gtk.Grid()
        grid_elec.add_css_class("details-grid")
        grid_elec.set_row_spacing(2)
        grid_elec.set_column_spacing(2)
        tab_elec.append(grid_elec)
        
        # Populate Elec Table
        grid_elec.attach(make_cell("Параметр", is_header=True), 0, 0, 1, 1)
        grid_elec.attach(make_cell("День", is_header=True), 1, 0, 1, 1)
        grid_elec.attach(make_cell("Ніч", is_header=True), 2, 0, 1, 1)
        
        grid_elec.attach(make_cell("Поточний показник"), 0, 1, 1, 1)
        grid_elec.attach(make_cell(f"{e_curr_d or 0:.0f} кВт"), 1, 1, 1, 1)
        grid_elec.attach(make_cell(f"{e_curr_n or 0:.0f} кВт"), 2, 1, 1, 1)
        
        grid_elec.attach(make_cell("Попередній показник"), 0, 2, 1, 1)
        grid_elec.attach(make_cell(f"{e_prev_d or 0:.0f} кВт"), 1, 2, 1, 1)
        grid_elec.attach(make_cell(f"{e_prev_n or 0:.0f} кВт"), 2, 2, 1, 1)
        
        e_diff_d = (e_curr_d - e_prev_d) if (e_curr_d is not None and e_prev_d is not None) else 0.0
        e_diff_n = (e_curr_n - e_prev_n) if (e_curr_n is not None and e_prev_n is not None) else 0.0
        grid_elec.attach(make_cell("Споживання (Різниця)"), 0, 3, 1, 1)
        grid_elec.attach(make_cell(f"{e_diff_d:.0f} кВт", is_bold=True), 1, 3, 1, 1)
        grid_elec.attach(make_cell(f"{e_diff_n:.0f} кВт", is_bold=True), 2, 3, 1, 1)
        
        grid_elec.attach(make_cell("Тариф"), 0, 4, 1, 1)
        grid_elec.attach(make_cell(f"{record.get('elec_day_tariff', 0.0):.2f} грн/кВт"), 1, 4, 1, 1)
        grid_elec.attach(make_cell(f"{record.get('elec_night_tariff', 0.0):.2f} грн/кВт"), 2, 4, 1, 1)
        
        grid_elec.attach(make_cell("Сума за фазою"), 0, 5, 1, 1)
        grid_elec.attach(make_cell(f"{e_diff_d * record.get('elec_day_tariff', 0.0):.2f} грн"), 1, 5, 1, 1)
        grid_elec.attach(make_cell(f"{e_diff_n * record.get('elec_night_tariff', 0.0):.2f} грн"), 2, 5, 1, 1)
        
        # Display mathematical formula card below grid
        formula_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=5)
        formula_box.set_margin_top(10)
        
        lbl_f_title = Gtk.Label(label="Математична формула розрахунку:")
        lbl_f_title.add_css_class("value-label")
        lbl_f_title.set_halign(Gtk.Align.START)
        formula_box.append(lbl_f_title)
        
        formula_text = f"({e_diff_d:.0f} * {record.get('elec_day_tariff', 0.0):.2f}) + ({e_diff_n:.0f} * {record.get('elec_night_tariff', 0.0):.2f}) = {record.get('elec_calculated', 0.0):.2f} грн"
        lbl_f_val = Gtk.Label(label=formula_text)
        lbl_f_val.add_css_class("details-formula")
        lbl_f_val.set_halign(Gtk.Align.START)
        formula_box.append(lbl_f_val)
        
        tab_elec.append(formula_box)
        notebook.append_page(tab_elec, Gtk.Label(label="Електрика"))

        # ----------------------------------------------------
        # TAB 3: Газ (Gas and Distribution details)
        # ----------------------------------------------------
        tab_gas = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        tab_gas.set_margin_start(10)
        tab_gas.set_margin_end(10)
        tab_gas.set_margin_top(15)
        tab_gas.set_margin_bottom(10)
        
        lbl_g_title = Gtk.Label(label="Деталі розрахунку за Газ (Споживання та Розподіл)")
        lbl_g_title.add_css_class("details-tab-title")
        lbl_g_title.set_halign(Gtk.Align.START)
        tab_gas.append(lbl_g_title)
        
        grid_gas = Gtk.Grid()
        grid_gas.add_css_class("details-grid")
        grid_gas.set_row_spacing(2)
        grid_gas.set_column_spacing(2)
        tab_gas.append(grid_gas)
        
        grid_gas.attach(make_cell("Показник / Параметр", is_header=True), 0, 0, 1, 1)
        grid_gas.attach(make_cell("Споживання газу", is_header=True), 1, 0, 1, 1)
        grid_gas.attach(make_cell("Розподіл газу (Транспортування)", is_header=True), 2, 0, 1, 1)
        
        grid_gas.attach(make_cell("Поточний показник / Обсяг"), 0, 1, 1, 1)
        grid_gas.attach(make_cell(f"{g_curr or 0:.2f} м³" if g_curr is not None else "-"), 1, 1, 1, 1)
        grid_gas.attach(make_cell(f"{record.get('gas_dist_volume', 0.0):.2f} м³"), 2, 1, 1, 1)
        
        grid_gas.attach(make_cell("Попередній показник"), 0, 2, 1, 1)
        grid_gas.attach(make_cell(f"{g_prev or 0:.2f} м³" if g_prev is not None else "-"), 1, 2, 1, 1)
        grid_gas.attach(make_cell("-"), 2, 2, 1, 1)
        
        g_diff = (g_curr - g_prev) if (g_curr is not None and g_prev is not None) else 0.0
        grid_gas.attach(make_cell("Розрахунковий обсяг"), 0, 3, 1, 1)
        grid_gas.attach(make_cell(f"{g_diff:.2f} м³", is_bold=True), 1, 3, 1, 1)
        grid_gas.attach(make_cell(f"{record.get('gas_dist_volume', 0.0):.2f} м³", is_bold=True), 2, 3, 1, 1)
        
        grid_gas.attach(make_cell("Тариф"), 0, 4, 1, 1)
        grid_gas.attach(make_cell(f"{record.get('gas_tariff', 0.0):.4f} грн/м³"), 1, 4, 1, 1)
        grid_gas.attach(make_cell(f"{record.get('gas_dist_tariff', 0.0):.4f} грн/м³"), 2, 4, 1, 1)
        
        grid_gas.attach(make_cell("Розрахована сума"), 0, 5, 1, 1)
        grid_gas.attach(make_cell(f"{record.get('gas_calculated', 0.0):.2f} грн", is_bold=True), 1, 5, 1, 1)
        grid_gas.attach(make_cell(f"{record.get('gas_dist_calculated', 0.0):.2f} грн", is_bold=True), 2, 5, 1, 1)
        
        # Mathematical Formulas Gas
        formula_box_g = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=5)
        formula_box_g.set_margin_top(10)
        
        lbl_fg_title = Gtk.Label(label="Математичні формули розрахунку:")
        lbl_fg_title.add_css_class("value-label")
        lbl_fg_title.set_halign(Gtk.Align.START)
        formula_box_g.append(lbl_fg_title)
        
        formula_gas_consumption = f"Споживання: {g_diff:.2f} м³ * {record.get('gas_tariff', 0.0):.4f} = {record.get('gas_calculated', 0.0):.2f} грн"
        lbl_fg_c = Gtk.Label(label=formula_gas_consumption)
        lbl_fg_c.add_css_class("details-formula")
        lbl_fg_c.set_halign(Gtk.Align.START)
        formula_box_g.append(lbl_fg_c)
        
        formula_gas_dist = f"Розподіл: {record.get('gas_dist_volume', 0.0):.2f} м³ * {record.get('gas_dist_tariff', 0.0):.4f} = {record.get('gas_dist_calculated', 0.0):.2f} грн"
        lbl_fg_d = Gtk.Label(label=formula_gas_dist)
        lbl_fg_d.add_css_class("details-formula")
        lbl_fg_d.set_halign(Gtk.Align.START)
        lbl_fg_d.set_margin_top(4)
        formula_box_g.append(lbl_fg_d)
        
        tab_gas.append(formula_box_g)
        notebook.append_page(tab_gas, Gtk.Label(label="Газ"))

        # ----------------------------------------------------
        # TAB 4: Водоканал (Water Details)
        # ----------------------------------------------------
        tab_water = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        tab_water.set_margin_start(10)
        tab_water.set_margin_end(10)
        tab_water.set_margin_top(15)
        tab_water.set_margin_bottom(10)
        
        lbl_w_title = Gtk.Label(label="Деталі розрахунку за Водопостачання та Водовідведення")
        lbl_w_title.add_css_class("details-tab-title")
        lbl_w_title.set_halign(Gtk.Align.START)
        tab_water.append(lbl_w_title)
        
        grid_water = Gtk.Grid()
        grid_water.add_css_class("details-grid")
        grid_water.set_row_spacing(2)
        grid_water.set_column_spacing(2)
        tab_water.append(grid_water)
        
        w_diff = (w_curr - w_prev) if (w_curr is not None and w_prev is not None) else 0.0
        
        grid_water.attach(make_cell("Параметр розрахунку", is_header=True), 0, 0, 1, 1)
        grid_water.attach(make_cell("Значення / Показник", is_header=True), 1, 0, 1, 1)
        grid_water.attach(make_cell("Формула розрахунку складової", is_header=True), 2, 0, 1, 1)
        
        grid_water.attach(make_cell("Поточний показник"), 0, 1, 1, 1)
        grid_water.attach(make_cell(f"{w_curr or 0:.2f} м³" if w_curr is not None else "-"), 1, 1, 1, 1)
        grid_water.attach(make_cell("-"), 2, 1, 1, 1)
        
        grid_water.attach(make_cell("Попередній показник"), 0, 2, 1, 1)
        grid_water.attach(make_cell(f"{w_prev or 0:.2f} м³" if w_prev is not None else "-"), 1, 2, 1, 1)
        grid_water.attach(make_cell("-"), 2, 2, 1, 1)
        
        grid_water.attach(make_cell("Спожита вода (Обсяг)"), 0, 3, 1, 1)
        grid_water.attach(make_cell(f"{w_diff:.2f} м³", is_bold=True), 1, 3, 1, 1)
        grid_water.attach(make_cell("-"), 2, 3, 1, 1)
        
        w_supply_cost = w_diff * record.get('water_supply_tariff', 0.0)
        grid_water.attach(make_cell("Водопостачання"), 0, 4, 1, 1)
        grid_water.attach(make_cell(f"{record.get('water_supply_tariff', 0.0):.2f} грн/м³"), 1, 4, 1, 1)
        grid_water.attach(make_cell(f"{w_diff:.2f} * {record.get('water_supply_tariff', 0.0):.2f} = {w_supply_cost:.2f} грн"), 2, 4, 1, 1)
        
        w_drainage_cost = w_diff * record.get('water_drainage_tariff', 0.0)
        grid_water.attach(make_cell("Водовідведення"), 0, 5, 1, 1)
        grid_water.attach(make_cell(f"{record.get('water_drainage_tariff', 0.0):.2f} грн/м³"), 1, 5, 1, 1)
        grid_water.attach(make_cell(f"{w_diff:.2f} * {record.get('water_drainage_tariff', 0.0):.2f} = {w_drainage_cost:.2f} грн"), 2, 5, 1, 1)
        
        grid_water.attach(make_cell("Абонентська плата"), 0, 6, 1, 1)
        grid_water.attach(make_cell(f"{record.get('water_sub_tariff', 0.0):.2f} грн"), 1, 6, 1, 1)
        grid_water.attach(make_cell("Фіксована абонплата"), 2, 6, 1, 1)
        
        grid_water.attach(make_cell("Загальна розрахована сума", is_bold=True), 0, 7, 1, 1)
        grid_water.attach(make_cell(f"{record.get('water_calculated', 0.0):.2f} грн", is_bold=True), 1, 7, 1, 1)
        grid_water.attach(make_cell("Сума трьох складових"), 2, 7, 1, 1)
        
        # Display mathematical formula water
        formula_box_w = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=5)
        formula_box_w.set_margin_top(10)
        
        lbl_fw_title = Gtk.Label(label="Математична формула розрахунку:")
        lbl_fw_title.add_css_class("value-label")
        lbl_fw_title.set_halign(Gtk.Align.START)
        formula_box_w.append(lbl_fw_title)
        
        formula_water_text = f"({w_diff:.2f} м³ * {record.get('water_supply_tariff', 0.0):.2f}) + ({w_diff:.2f} м³ * {record.get('water_drainage_tariff', 0.0):.2f}) + {record.get('water_sub_tariff', 0.0):.2f} = {record.get('water_calculated', 0.0):.2f} грн"
        lbl_fw_val = Gtk.Label(label=formula_water_text)
        lbl_fw_val.add_css_class("details-formula")
        lbl_fw_val.set_halign(Gtk.Align.START)
        formula_box_w.append(lbl_fw_val)
        
        tab_water.append(formula_box_w)
        notebook.append_page(tab_water, Gtk.Label(label="Водоканал"))

        # ----------------------------------------------------
        # TAB 5: Вивіз сміття (Garbage)
        # ----------------------------------------------------
        tab_garbage = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        tab_garbage.set_margin_start(10)
        tab_garbage.set_margin_end(10)
        tab_garbage.set_margin_top(15)
        tab_garbage.set_margin_bottom(10)
        
        lbl_m_title = Gtk.Label(label="Деталі розрахунку за вивіз сміття")
        lbl_m_title.add_css_class("details-tab-title")
        lbl_m_title.set_halign(Gtk.Align.START)
        tab_garbage.append(lbl_m_title)
        
        grid_garbage = Gtk.Grid()
        grid_garbage.add_css_class("details-grid")
        grid_garbage.set_row_spacing(2)
        grid_garbage.set_column_spacing(2)
        tab_garbage.append(grid_garbage)
        
        grid_garbage.attach(make_cell("Параметр розрахунку", is_header=True), 0, 0, 1, 1)
        grid_garbage.attach(make_cell("Значення / Показник", is_header=True), 1, 0, 1, 1)
        
        grid_garbage.attach(make_cell("Тип оплати"), 0, 1, 1, 1)
        grid_garbage.attach(make_cell("Фіксований тариф (без лічильника)"), 1, 1, 1, 1)
        
        grid_garbage.attach(make_cell("Діючий тариф за місяць"), 0, 2, 1, 1)
        grid_garbage.attach(make_cell(f"{record.get('garbage_tariff', 0.0):.2f} грн"), 1, 2, 1, 1)
        
        grid_garbage.attach(make_cell("Загальна розрахована сума", is_bold=True), 0, 3, 1, 1)
        grid_garbage.attach(make_cell(f"{record.get('garbage_calculated', 0.0):.2f} грн", is_bold=True), 1, 3, 1, 1)
        
        notebook.append_page(tab_garbage, Gtk.Label(label="Сміття"))

    def on_export_clicked(self, btn):
        all_readings = db.get_all_readings()
        if not all_readings:
            self.show_error_dialog("Помилка", "Немає записів для експорту!")
            return
            
        periods = sorted([r['period'] for r in all_readings])
        
        dialog = ModalWindow(title="Налаштування експорту", transient_for=self.win, modal=True)
        dialog.set_default_size(350, -1)
        
        content = dialog.get_content_area()
        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=12)
        box.set_margin_start(15)
        box.set_margin_end(15)
        box.set_margin_top(15)
        box.set_margin_bottom(15)
        content.append(box)
        
        lbl_msg = Gtk.Label(label="Виберіть діапазон періодів для експорту:")
        lbl_msg.set_halign(Gtk.Align.START)
        box.append(lbl_msg)
        
        grid = Gtk.Grid()
        grid.set_row_spacing(8)
        grid.set_column_spacing(10)
        box.append(grid)
        
        lbl_start = Gtk.Label(label="Початковий місяць:")
        lbl_start.set_halign(Gtk.Align.START)
        grid.attach(lbl_start, 0, 0, 1, 1)
        
        combo_start = create_string_dropdown(periods)
        grid.attach(combo_start, 1, 0, 1, 1)
        
        lbl_end = Gtk.Label(label="Кінцевий місяць:")
        lbl_end.set_halign(Gtk.Align.START)
        grid.attach(lbl_end, 0, 1, 1, 1)
        
        combo_end = create_string_dropdown(periods)
        grid.attach(combo_end, 1, 1, 1, 1)
        
        # Set defaults: last 12 months range
        default_start_idx = max(0, len(periods) - 12)
        default_end_idx = len(periods) - 1
        combo_start.set_selected(default_start_idx)
        combo_end.set_selected(default_end_idx)
        
        lbl_format = Gtk.Label(label="Формат файлу: Excel Spreadsheet (.xlsx)")
        lbl_format.set_halign(Gtk.Align.START)
        lbl_format.add_css_class("value-label")
        box.append(lbl_format)
        
        # Action buttons
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_halign(Gtk.Align.START)
        btn_box.set_margin_top(4)
        box.append(btn_box)
        
        btn_export = Gtk.Button(label="Експортувати")
        btn_export.add_css_class("btn-primary")
        btn_box.append(btn_export)
        
        btn_cancel = Gtk.Button(label="Скасувати")
        btn_cancel.add_css_class("btn-secondary")
        btn_box.append(btn_cancel)
        
        def do_export(b):
            start_item = combo_start.get_selected_item()
            end_item = combo_end.get_selected_item()
            start_p = start_item.get_string() if start_item else None
            end_p = end_item.get_string() if end_item else None
            
            if not start_p or not end_p:
                return
            if start_p > end_p:
                self.show_error_dialog("Помилка", "Початковий місяць не може бути пізнішим за кінцевий!")
                return
                
            dialog.destroy()
            self.run_export_file_chooser(start_p, end_p)
            
        btn_export.connect("clicked", do_export)
        btn_cancel.connect("clicked", lambda x: dialog.destroy())
        
        dialog.present()

    def run_export_file_chooser(self, start_period, end_period):
        """Opens a native Save file dialog using the modern Gtk.FileDialog API
        (replaces the deprecated Gtk.FileChooserDialog).
        """
        file_dialog = Gtk.FileDialog()
        file_dialog.set_title("Зберегти звіт Excel")

        # Add .xlsx filter
        filters = Gio.ListStore.new(Gtk.FileFilter)
        xlsx_filter = Gtk.FileFilter()
        xlsx_filter.set_name("Excel файли (*.xlsx)")
        xlsx_filter.add_suffix("xlsx")
        filters.append(xlsx_filter)
        file_dialog.set_filters(filters)

        # Suggest filename
        default_name = f"Комуналка_експорт_{start_period}_to_{end_period}.xlsx"
        file_dialog.set_initial_name(default_name)

        def on_save_done(dialog, result):
            try:
                gfile = dialog.save_finish(result)
                if gfile:
                    file_path = gfile.get_path()
                    if not file_path.lower().endswith(".xlsx"):
                        file_path += ".xlsx"
                    self.export_to_excel(start_period, end_period, file_path)
            except GLib.Error:
                pass  # User cancelled

        file_dialog.save(self.win, None, on_save_done)


    def export_to_excel(self, start_period, end_period, filepath):
        try:
            wb = openpyxl.Workbook()
            
            readings = db.get_readings_in_range(start_period, end_period)
            if not readings:
                self.show_error_dialog("Помилка", "Не знайдено даних у вибраному діапазоні!")
                return
                
            # Styling config
            font_title = Font(name='Segoe UI', size=14, bold=True, color='1F4E79')
            font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            font_body = Font(name='Segoe UI', size=10)
            font_bold = Font(name='Segoe UI', size=10, bold=True)
            
            fill_header = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            fill_stripe = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
            fill_total = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            double_bottom = Border(
                bottom=Side(style='double', color='000000'),
                top=Side(style='thin', color='D9D9D9')
            )
            
            align_left = Alignment(horizontal='left', vertical='center')
            align_right = Alignment(horizontal='right', vertical='center')
            align_center = Alignment(horizontal='center', vertical='center')

            def get_status_text(status):
                return STATUS_LABELS.get(status, STATUS_LABELS[0])

            # ----------------------------------------------------
            # SHEET 1: Зведена (Summary Sheet)
            # ----------------------------------------------------
            ws_sum = wb.active
            ws_sum.title = "Зведена"
            
            ws_sum.append(["Зведена таблиця розрахунків комунальних послуг"])
            ws_sum.cell(row=1, column=1).font = font_title
            ws_sum.row_dimensions[1].height = 30
            
            ws_sum.append([f"Діапазон періодів: {start_period} — {end_period}"])
            ws_sum.append([]) # Empty spacer
            
            headers = [
                "Період", "Електрика Розраховано", "Електрика Сплачено", 
                "Газ Споживання Розраховано", "Газ Споживання Сплачено",
                "Газ Розподіл Розраховано", "Газ Розподіл Сплачено",
                "Водоканал Розраховано", "Водоканал Сплачено",
                "Вивіз сміття Розраховано", "Вивіз сміття Сплачено",
                "РАЗОМ Розраховано", "РАЗОМ Сплачено"
            ]
            ws_sum.append(headers)
            ws_sum.row_dimensions[4].height = 24
            
            for col_idx, h in enumerate(headers, start=1):
                cell = ws_sum.cell(row=4, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            for row_idx, r in enumerate(readings, start=5):
                period = r['period']
                
                e_calc = r.get('elec_calculated', 0.0) or 0.0
                e_paid = r.get('elec_paid', 0.0) or 0.0 if r.get('elec_paid_status', 0) == 1 else 0.0
                
                g_calc = r.get('gas_calculated', 0.0) or 0.0
                g_paid = r.get('gas_paid', 0.0) or 0.0 if r.get('gas_paid_status', 0) == 1 else 0.0
                
                gd_calc = r.get('gas_dist_calculated', 0.0) or 0.0
                gd_paid = r.get('gas_dist_paid', 0.0) or 0.0 if r.get('gas_dist_paid_status', 0) == 1 else 0.0
                
                w_calc = r.get('water_calculated', 0.0) or 0.0
                w_paid = r.get('water_paid', 0.0) or 0.0 if r.get('water_paid_status', 0) == 1 else 0.0
                
                gb_calc = r.get('garbage_calculated', 0.0) or 0.0
                gb_paid = r.get('garbage_paid', 0.0) or 0.0 if r.get('garbage_paid_status', 0) == 1 else 0.0
                
                tot_calc = e_calc + g_calc + gd_calc + w_calc + gb_calc
                tot_paid = e_paid + g_paid + gd_paid + w_paid + gb_paid
                
                row_data = [
                    period, e_calc, e_paid, g_calc, g_paid, gd_calc, gd_paid,
                    w_calc, w_paid, gb_calc, gb_paid, tot_calc, tot_paid
                ]
                ws_sum.append(row_data)
                ws_sum.row_dimensions[row_idx].height = 20
                
                is_stripe = (row_idx % 2 == 0)
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_sum.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = thin_border
                    if is_stripe:
                        cell.fill = fill_stripe
                    if col_idx == 1:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.00" грн"'
                        
            tot_row_idx = len(readings) + 5
            tot_cells = [
                "РАЗОМ",
                f"=SUM(B5:B{tot_row_idx-1})", f"=SUM(C5:C{tot_row_idx-1})",
                f"=SUM(D5:D{tot_row_idx-1})", f"=SUM(E5:E{tot_row_idx-1})",
                f"=SUM(F5:F{tot_row_idx-1})", f"=SUM(G5:G{tot_row_idx-1})",
                f"=SUM(H5:H{tot_row_idx-1})", f"=SUM(I5:I{tot_row_idx-1})",
                f"=SUM(J5:J{tot_row_idx-1})", f"=SUM(K5:K{tot_row_idx-1})",
                f"=SUM(L5:L{tot_row_idx-1})", f"=SUM(M5:M{tot_row_idx-1})"
            ]
            ws_sum.append(tot_cells)
            ws_sum.row_dimensions[tot_row_idx].height = 22
            
            for col_idx, val in enumerate(tot_cells, start=1):
                cell = ws_sum.cell(row=tot_row_idx, column=col_idx)
                cell.font = font_bold
                cell.fill = fill_total
                cell.border = double_bottom
                if col_idx == 1:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right
                    cell.number_format = '#,##0.00" грн"'

            for col in ws_sum.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # ----------------------------------------------------
            # SHEET 2: Електрика (Electricity Sheet)
            # ----------------------------------------------------
            ws_elec = wb.create_sheet("Електрика")
            ws_elec.append(["Детальний звіт за електроенергію"])
            ws_elec.cell(row=1, column=1).font = font_title
            ws_elec.row_dimensions[1].height = 30
            ws_elec.append([])
            
            e_headers = [
                "Період", "Поточний День", "Поточний Ніч", "Попередній День", "Попередній Ніч",
                "Спожито День (кВт)", "Спожито Ніч (кВт)", "Тариф День", "Тариф Ніч",
                "Розраховано", "Нараховано", "Сплачено", "Статус", "Дата оплати"
            ]
            ws_elec.append(e_headers)
            ws_elec.row_dimensions[3].height = 24
            
            for col_idx, h in enumerate(e_headers, start=1):
                cell = ws_elec.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            for row_idx, r in enumerate(readings, start=4):
                period = r['period']
                prev_r = db.get_previous_reading(period) or {}
                
                curr_d = r.get('elec_day_reading')
                prev_d = prev_r.get('elec_day_reading', curr_d)
                curr_n = r.get('elec_night_reading')
                prev_n = prev_r.get('elec_night_reading', curr_n)
                
                diff_d = (curr_d - prev_d) if (curr_d is not None and prev_d is not None) else 0.0
                diff_n = (curr_n - prev_n) if (curr_n is not None and prev_n is not None) else 0.0
                
                row_data = [
                    period, curr_d, curr_n, prev_d, prev_n, diff_d, diff_n,
                    r.get('elec_day_tariff', 0.0), r.get('elec_night_tariff', 0.0),
                    r.get('elec_calculated', 0.0), r.get('elec_billed'),
                    r.get('elec_paid', 0.0) if r.get('elec_paid_status', 0) == 1 else 0.0,
                    get_status_text(r.get('elec_paid_status', 0)),
                    r.get('elec_paid_date', '')
                ]
                ws_elec.append(row_data)
                ws_elec.row_dimensions[row_idx].height = 20
                
                is_stripe = (row_idx % 2 == 0)
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_elec.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = thin_border
                    if is_stripe:
                        cell.fill = fill_stripe
                    if col_idx in [1, 13, 14]:
                        cell.alignment = align_center
                    elif col_idx in [2, 3, 4, 5, 6, 7]:
                        cell.alignment = align_right
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.00'
                        
            for col in ws_elec.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws_elec.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # ----------------------------------------------------
            # SHEET 3: Газ (Gas Sheet)
            # ----------------------------------------------------
            ws_gas = wb.create_sheet("Газ")
            ws_gas.append(["Детальний звіт за Газ (Споживання та Розподіл)"])
            ws_gas.cell(row=1, column=1).font = font_title
            ws_gas.row_dimensions[1].height = 30
            ws_gas.append([])
            
            g_headers = [
                "Період", "Поточний Газ", "Попередній Газ", "Спожито (м³)", "Тариф Газ", "Споживання Розраховано",
                "Споживання Нараховано", "Споживання Сплачено", "Споживання Статус",
                "Обсяг Розподілу", "Тариф Розподілу", "Розподіл Розраховано", "Розподіл Нараховано",
                "Розподіл Сплачено", "Розподіл Статус"
            ]
            ws_gas.append(g_headers)
            ws_gas.row_dimensions[3].height = 24
            
            for col_idx, h in enumerate(g_headers, start=1):
                cell = ws_gas.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            for row_idx, r in enumerate(readings, start=4):
                period = r['period']
                prev_r = db.get_previous_reading(period) or {}
                
                curr_g = r.get('gas_reading')
                prev_g = prev_r.get('gas_reading', curr_g)
                diff_g = (curr_g - prev_g) if (curr_g is not None and prev_g is not None) else 0.0
                
                row_data = [
                    period, curr_g, prev_g, diff_g, r.get('gas_tariff', 0.0),
                    r.get('gas_calculated', 0.0), r.get('gas_billed'),
                    r.get('gas_paid', 0.0) if r.get('gas_paid_status', 0) == 1 else 0.0,
                    get_status_text(r.get('gas_paid_status', 0)),
                    r.get('gas_dist_volume', 0.0), r.get('gas_dist_tariff', 0.0),
                    r.get('gas_dist_calculated', 0.0), r.get('gas_dist_billed'),
                    r.get('gas_dist_paid', 0.0) if r.get('gas_dist_paid_status', 0) == 1 else 0.0,
                    get_status_text(r.get('gas_dist_paid_status', 0))
                ]
                ws_gas.append(row_data)
                ws_gas.row_dimensions[row_idx].height = 20
                
                is_stripe = (row_idx % 2 == 0)
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_gas.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = thin_border
                    if is_stripe:
                        cell.fill = fill_stripe
                    if col_idx in [1, 9, 15]:
                        cell.alignment = align_center
                    elif col_idx in [2, 3, 4]:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.00'
                    elif col_idx in [5, 11]:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.0000'
                    else:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.00'
                        
            for col in ws_gas.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws_gas.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # ----------------------------------------------------
            # SHEET 4: Вода (Water Sheet)
            # ----------------------------------------------------
            ws_water = wb.create_sheet("Вода")
            ws_water.append(["Детальний звіт за Водопостачання та Водовідведення"])
            ws_water.cell(row=1, column=1).font = font_title
            ws_water.row_dimensions[1].height = 30
            ws_water.append([])
            
            w_headers = [
                "Період", "Поточний Вода", "Попередній Вода", "Спожито (м³)", "Тариф Постач", "Тариф Водовідв",
                "Абонплата", "Розраховано", "Нараховано", "Сплачено", "Статус"
            ]
            ws_water.append(w_headers)
            ws_water.row_dimensions[3].height = 24
            
            for col_idx, h in enumerate(w_headers, start=1):
                cell = ws_water.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            for row_idx, r in enumerate(readings, start=4):
                period = r['period']
                prev_r = db.get_previous_reading(period) or {}
                
                curr_w = r.get('water_reading')
                prev_w = prev_r.get('water_reading', curr_w)
                diff_w = (curr_w - prev_w) if (curr_w is not None and prev_w is not None) else 0.0
                
                row_data = [
                    period, curr_w, prev_w, diff_w, r.get('water_supply_tariff', 0.0),
                    r.get('water_drainage_tariff', 0.0), r.get('water_sub_tariff', 0.0),
                    r.get('water_calculated', 0.0), r.get('water_billed'),
                    r.get('water_paid', 0.0) if r.get('water_paid_status', 0) == 1 else 0.0,
                    get_status_text(r.get('water_paid_status', 0))
                ]
                ws_water.append(row_data)
                ws_water.row_dimensions[row_idx].height = 20
                
                is_stripe = (row_idx % 2 == 0)
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_water.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = thin_border
                    if is_stripe:
                        cell.fill = fill_stripe
                    if col_idx in [1, 11]:
                        cell.alignment = align_center
                    elif col_idx in [2, 3, 4]:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.00'
                        
            for col in ws_water.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws_water.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # ----------------------------------------------------
            # SHEET 5: Сміття (Garbage Sheet)
            # ----------------------------------------------------
            ws_garbage = wb.create_sheet("Сміття")
            ws_garbage.append(["Детальний звіт за вивіз сміття"])
            ws_garbage.cell(row=1, column=1).font = font_title
            ws_garbage.row_dimensions[1].height = 30
            ws_garbage.append([])
            
            m_headers = [
                "Період", "Діючий Тариф за місяць", "Розраховано", "Нараховано", "Сплачено", "Статус"
            ]
            ws_garbage.append(m_headers)
            ws_garbage.row_dimensions[3].height = 24
            
            for col_idx, h in enumerate(m_headers, start=1):
                cell = ws_garbage.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            for row_idx, r in enumerate(readings, start=4):
                period = r['period']
                
                row_data = [
                    period, r.get('garbage_tariff', 0.0), r.get('garbage_calculated', 0.0),
                    r.get('garbage_billed'),
                    r.get('garbage_paid', 0.0) if r.get('garbage_paid_status', 0) == 1 else 0.0,
                    get_status_text(r.get('garbage_paid_status', 0))
                ]
                ws_garbage.append(row_data)
                ws_garbage.row_dimensions[row_idx].height = 20
                
                is_stripe = (row_idx % 2 == 0)
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_garbage.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = thin_border
                    if is_stripe:
                        cell.fill = fill_stripe
                    if col_idx in [1, 6]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_right
                        cell.number_format = '#,##0.00'
                        
            for col in ws_garbage.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws_garbage.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(filepath)
            self.show_info_dialog("Успішно", f"Звіт успішно експортовано в файл:\n{filepath}")
        except Exception as e:
            self.show_error_dialog("Помилка експорту", f"Не вдалося виконати експорт:\n{str(e)}")

    def build_edit_history_view(self):
        scroll = Gtk.ScrolledWindow()
        scroll.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        
        self.edit_hist_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=15)
        self.edit_hist_box.add_css_class("dashboard-container")
        scroll.set_child(self.edit_hist_box)
        
        self.stack.add_named(scroll, "edit_history")

    def show_edit_history_view(self, period):
        # Switch tab directly and deactivate sidebar buttons
        self.active_tab = "edit_history"
        self.stack.set_visible_child_name("edit_history")
        
        self.btn_dash.remove_css_class("sidebar-button-active")
        self.btn_hist.remove_css_class("sidebar-button-active")
        self.btn_tariffs.remove_css_class("sidebar-button-active")
        self.btn_accs.remove_css_class("sidebar-button-active")
        
        # Clear previous elements from container
        self._clear_container(self.edit_hist_box)
            
        record = db.get_reading(period)
        if not record:
            self.show_error_dialog("Помилка", f"Не знайдено даних для періоду {period}!")
            self.switch_tab("history")
            return
            
        # Format month name nicely using shared helper
        period_name = self._format_period_name(period)
            
        # Header navigation row
        header_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=15)
        header_box.add_css_class("details-header-box")
        
        btn_back = Gtk.Button(label="← Скасувати")
        btn_back.add_css_class("btn-secondary")
        btn_back.connect("clicked", lambda x: self.switch_tab("history"))
        header_box.append(btn_back)
        
        lbl_title = Gtk.Label(label=f"Редагування періоду: {period_name}")
        lbl_title.add_css_class("dashboard-section-title")
        lbl_title.set_hexpand(True)
        lbl_title.set_halign(Gtk.Align.START)
        header_box.append(lbl_title)
        
        self.edit_hist_box.append(header_box)
        
        # Red Warning Banner
        banner = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=5)
        banner.add_css_class("reminder-banner")
        
        lbl_w_title = Gtk.Label(label=f"⚠️ Увага! Ви редагуєте архівний запис за період {period_name}.")
        lbl_w_title.set_halign(Gtk.Align.START)
        lbl_w_title.add_css_class("banner-title")
        banner.append(lbl_w_title)
        
        lbl_w_desc = Gtk.Label(label="Будь ласка, будьте обережні. Зміни змінять зафіксовані історичні розрахунки. Рекомендуємо зробити резервну копію за допомогою кнопки Експорт перед початком редагування.")
        lbl_w_desc.set_halign(Gtk.Align.START)
        lbl_w_desc.set_wrap(True)
        banner.append(lbl_w_desc)
        self.edit_hist_box.append(banner)
        
        # Grid of Cards
        grid = Gtk.Grid()
        grid.set_row_spacing(10)
        grid.set_column_spacing(10)
        grid.set_row_homogeneous(True)
        grid.set_column_homogeneous(True)
        self.edit_hist_box.append(grid)
        
        def create_card(title):
            card = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=8)
            card.add_css_class("card")
            
            lbl = Gtk.Label(label=title)
            lbl.add_css_class("card-title")
            lbl.set_halign(Gtk.Align.START)
            card.append(lbl)
            
            card_grid = Gtk.Grid()
            card_grid.set_row_spacing(6)
            card_grid.set_column_spacing(10)
            card_grid.set_column_homogeneous(True)
            card.append(card_grid)
            return card, card_grid
            
        def add_field(card_grid, row_idx, label_text, val):
            lbl = Gtk.Label(label=label_text)
            lbl.set_halign(Gtk.Align.START)
            lbl.add_css_class("value-label")
            card_grid.attach(lbl, 0, row_idx, 1, 1)
            
            entry = Gtk.Entry()
            entry.set_text(str(val) if val is not None else "")
            entry.set_hexpand(True)
            card_grid.attach(entry, 1, row_idx, 1, 1)
            return entry

        def add_status_combo(card_grid, row_idx, val):
            lbl = Gtk.Label(label="Статус платежу:")
            lbl.set_halign(Gtk.Align.START)
            lbl.add_css_class("value-label")
            card_grid.attach(lbl, 0, row_idx, 1, 1)
            
            combo = create_string_dropdown(["Не сплачено", "Сплачено", "Аванс / Пропуск"], val)
            combo.set_hexpand(True)
            card_grid.attach(combo, 1, row_idx, 1, 1)
            return combo

        # 1. Electricity Card
        card_e, grid_e = create_card("Електрика")
        ent_e_day = add_field(grid_e, 0, "День Показник (кВт):", record.get('elec_day_reading'))
        ent_e_night = add_field(grid_e, 1, "Ніч Показник (кВт):", record.get('elec_night_reading'))
        ent_e_day_t = add_field(grid_e, 2, "День Тариф (грн):", record.get('elec_day_tariff'))
        ent_e_night_t = add_field(grid_e, 3, "Ніч Тариф (грн):", record.get('elec_night_tariff'))
        ent_e_billed = add_field(grid_e, 4, "Нараховано (грн):", record.get('elec_billed'))
        ent_e_paid = add_field(grid_e, 5, "Сплачено (грн):", record.get('elec_paid'))
        combo_e_status = add_status_combo(grid_e, 6, record.get('elec_paid_status', 0))
        grid.attach(card_e, 0, 0, 1, 1)
        
        # 2. Gas Consumption Card
        card_g, grid_g = create_card("Газ (Споживання)")
        ent_g = add_field(grid_g, 0, "Показник лічильника (м³):", record.get('gas_reading'))
        ent_g_t = add_field(grid_g, 1, "Тариф (грн):", record.get('gas_tariff'))
        ent_g_billed = add_field(grid_g, 2, "Нараховано (грн):", record.get('gas_billed'))
        ent_g_paid = add_field(grid_g, 3, "Сплачено (грн):", record.get('gas_paid'))
        combo_g_status = add_status_combo(grid_g, 4, record.get('gas_paid_status', 0))
        grid.attach(card_g, 1, 0, 1, 1)
        
        # 3. Gas Distribution Card
        card_gd, grid_gd = create_card("Газ (Розподіл)")
        ent_gd_v = add_field(grid_gd, 0, "Обсяг розподілу (м³):", record.get('gas_dist_volume'))
        ent_gd_t = add_field(grid_gd, 1, "Тариф розподілу (грн):", record.get('gas_dist_tariff'))
        ent_gd_billed = add_field(grid_gd, 2, "Нараховано (грн):", record.get('gas_dist_billed'))
        ent_gd_paid = add_field(grid_gd, 3, "Сплачено (грн):", record.get('gas_dist_paid'))
        combo_gd_status = add_status_combo(grid_gd, 4, record.get('gas_dist_paid_status', 0))
        grid.attach(card_gd, 0, 1, 1, 1)
        
        # 4. Water Card
        card_w, grid_w = create_card("Водоканал")
        ent_w = add_field(grid_w, 0, "Показник лічильника (м³):", record.get('water_reading'))
        ent_w_t_s = add_field(grid_w, 1, "Тариф постачання (грн):", record.get('water_supply_tariff'))
        ent_w_t_d = add_field(grid_w, 2, "Тариф водовідведення (грн):", record.get('water_drainage_tariff'))
        ent_w_t_sub = add_field(grid_w, 3, "Абонентська плата (грн):", record.get('water_sub_tariff'))
        ent_w_billed = add_field(grid_w, 4, "Нараховано (грн):", record.get('water_billed'))
        ent_w_paid = add_field(grid_w, 5, "Сплачено (грн):", record.get('water_paid'))
        combo_w_status = add_status_combo(grid_w, 6, record.get('water_paid_status', 0))
        grid.attach(card_w, 1, 1, 1, 1)
        
        # 5. Garbage Card
        card_m, grid_m = create_card("Вивіз сміття")
        ent_m_t = add_field(grid_m, 0, "Фіксований тариф (грн):", record.get('garbage_tariff'))
        ent_m_billed = add_field(grid_m, 1, "Нараховано (грн):", record.get('garbage_billed'))
        ent_m_paid = add_field(grid_m, 2, "Сплачено (грн):", record.get('garbage_paid'))
        combo_m_status = add_status_combo(grid_m, 3, record.get('garbage_paid_status', 0))
        grid.attach(card_m, 0, 2, 1, 1)
        
        # Bottom Save Box
        btn_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        btn_box.set_margin_top(15)
        
        btn_save = Gtk.Button(label="Зберегти зміни")
        btn_save.add_css_class("btn-primary")
        btn_box.append(btn_save)
        
        btn_cancel = Gtk.Button(label="Скасувати")
        btn_cancel.add_css_class("btn-secondary")
        btn_box.append(btn_cancel)
        
        self.edit_hist_box.append(btn_box)
        
        def get_val(entry, name):
            txt = entry.get_text().strip()
            if not txt:
                return None
            try:
                return float(txt)
            except ValueError:
                raise ValueError(f"Поле '{name}' повинно бути числовим значенням!")
                
        def do_save(btn):
            try:
                e_day = get_val(ent_e_day, "Електрика День Показник")
                e_night = get_val(ent_e_night, "Електрика Ніч Показник")
                e_day_t = get_val(ent_e_day_t, "Електрика День Тариф")
                e_night_t = get_val(ent_e_night_t, "Електрика Ніч Тариф")
                e_billed = get_val(ent_e_billed, "Електрика Нараховано")
                e_paid = get_val(ent_e_paid, "Електрика Сплачено")
                e_status = combo_e_status.get_selected()
                
                gas_val = get_val(ent_g, "Газ Показник")
                gas_t = get_val(ent_g_t, "Газ Тариф")
                gas_billed = get_val(ent_g_billed, "Газ Нараховано")
                gas_paid = get_val(ent_g_paid, "Газ Сплачено")
                gas_status = combo_g_status.get_selected()
                
                gd_vol = get_val(ent_gd_v, "Газ Розподіл Обсяг")
                gd_t = get_val(ent_gd_t, "Газ Розподіл Тариф")
                gd_billed = get_val(ent_gd_billed, "Газ Розподіл Нараховано")
                gd_paid = get_val(ent_gd_paid, "Газ Розподіл Сплачено")
                gd_status = combo_gd_status.get_selected()
                
                water_val = get_val(ent_w, "Вода Показник")
                w_t_s = get_val(ent_w_t_s, "Вода Тариф постачання")
                w_t_d = get_val(ent_w_t_d, "Вода Тариф водовідведення")
                w_t_sub = get_val(ent_w_t_sub, "Вода Абонплата")
                w_billed = get_val(ent_w_billed, "Вода Нараховано")
                w_paid = get_val(ent_w_paid, "Вода Сплачено")
                w_status = combo_w_status.get_selected()
                
                garbage_t = get_val(ent_m_t, "Сміття Тариф")
                garbage_billed = get_val(ent_m_billed, "Сміття Нараховано")
                garbage_paid = get_val(ent_m_paid, "Сміття Сплачено")
                garbage_status = combo_m_status.get_selected()
                
                reading_data = {
                    'period': period,
                    
                    'elec_day_reading': e_day,
                    'elec_night_reading': e_night,
                    'elec_day_tariff': e_day_t,
                    'elec_night_tariff': e_night_t,
                    'elec_billed': e_billed,
                    'elec_paid': e_paid,
                    'elec_paid_status': e_status,
                    'elec_paid_date': record.get('elec_paid_date') if e_status == record.get('elec_paid_status') else (datetime.now().strftime("%Y-%m-%d") if e_status == 1 else None),
                    
                    'gas_reading': gas_val,
                    'gas_tariff': gas_t,
                    'gas_billed': gas_billed,
                    'gas_paid': gas_paid,
                    'gas_paid_status': gas_status,
                    'gas_paid_date': record.get('gas_paid_date') if gas_status == record.get('gas_paid_status') else (datetime.now().strftime("%Y-%m-%d") if gas_status == 1 else None),
                    
                    'gas_dist_volume': gd_vol,
                    'gas_dist_tariff': gd_t,
                    'gas_dist_billed': gd_billed,
                    'gas_dist_paid': gd_paid,
                    'gas_dist_paid_status': gd_status,
                    'gas_dist_paid_date': record.get('gas_dist_paid_date') if gd_status == record.get('gas_dist_paid_status') else (datetime.now().strftime("%Y-%m-%d") if gd_status == 1 else None),
                    
                    'water_reading': water_val,
                    'water_supply_tariff': w_t_s,
                    'water_drainage_tariff': w_t_d,
                    'water_sub_tariff': w_t_sub,
                    'water_billed': w_billed,
                    'water_paid': w_paid,
                    'water_paid_status': w_status,
                    'water_paid_date': record.get('water_paid_date') if w_status == record.get('water_paid_status') else (datetime.now().strftime("%Y-%m-%d") if w_status == 1 else None),
                    
                    'garbage_tariff': garbage_t,
                    'garbage_billed': garbage_billed,
                    'garbage_paid': garbage_paid,
                    'garbage_paid_status': garbage_status,
                    'garbage_paid_date': record.get('garbage_paid_date') if garbage_status == record.get('garbage_paid_status') else (datetime.now().strftime("%Y-%m-%d") if garbage_status == 1 else None)
                }
                
                # Perform recalculations
                reading_data = UtilityCalculator.calculate_all(period, reading_data)
                
                # Save and refresh
                db.save_reading(reading_data)
                self.refresh_dashboard()
                self.refresh_history()
                
                self.show_info_dialog("Успішно", f"Зміни для періоду {period} збережено та перераховано!")
                self.switch_tab("history")
            except ValueError as e:
                self.show_error_dialog("Помилка вводу", str(e))
                
        btn_save.connect("clicked", do_save)
        btn_cancel.connect("clicked", lambda x: self.switch_tab("history"))

if __name__ == "__main__":
    GLib.set_prgname('io.github.f0ska.utilitytracker')
    GLib.set_application_name('UtilityTracker')
    app = UtilityTrackerApp()
    sys.exit(app.run(sys.argv))
